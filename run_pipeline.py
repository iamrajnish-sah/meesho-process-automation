#!/usr/bin/env python3
"""
run_pipeline.py — orchestrates all six sheet modules in dependency order.

Pipeline: raw_working_sheet → raw_data → meesho_prelim → error_margin
          → fk_prelim → client_prelim

Each step returns an anchor contract passed explicitly to the next step.
Step 7 updates New_Meesho Masterfile (formula drag + hardcoded carry-forward).

Usage:
    python run_pipeline.py "MeeshoMasterfilesampleautomation.xlsx" \\
        --month "May'26" --raw-data raw_may26.csv [--dry-run]
"""
import argparse
import os
import shutil
import sys
import tempfile
from typing import Dict, Literal, Optional, Set, Tuple

from openpyxl import load_workbook

import client_prelim
import error_margin
import fk_prelim
import meesho_prelim
import new_meesho_masterfile
import raw_data
import raw_working_sheet
from shared.fk_manual_inputs import FKManualInputs
from shared.hide_month_columns import (
    hide_month_columns_interactive,
    parse_visible_month_labels,
)
from shared.raw_input import load_raw_data
from shared.month_utils import MonthSequenceError, next_month_label_after
from shared.workbook_io import load_workbook_safe

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass


class PipelineValidationError(Exception):
    """Raised when GMV validation fails (used by web UI instead of sys.exit)."""


PipelineStage = Literal["raw_data", "error_margin", "client_prelim", "full_master"]

STAGE_STOP_AFTER = {
    "raw_data": 2,
    "error_margin": 4,
    "client_prelim": 6,
    "full_master": 7,
}

FINAL_OUTPUT_FILENAME = "meesho final output check.xlsx"


def validate_gmv_totals(
    raw_data_dict: Dict[str, Tuple[int, int]],
    matched_by_name: Dict[str, Tuple[int, int]],
    *,
    raise_error: bool = False,
) -> None:
    raw_total     = sum(v[1] for v in raw_data_dict.values())
    written_total = sum(v[1] for v in matched_by_name.values())
    raw_cr     = raw_total     / 1e7
    written_cr = written_total / 1e7
    print(f"\n  [VALIDATE] Raw input total GMV : {raw_cr:>12,.2f} INR Cr  ({raw_total:,})")
    print(f"  [VALIDATE] Written to RWS total: {written_cr:>12,.2f} INR Cr  ({written_total:,})")
    if raw_total != written_total:
        diff_cr = (raw_total - written_total) / 1e7
        msg = (
            f"VALIDATION FAILED — GMV totals do not match!\n"
            f"  Raw input:  {raw_cr:,.2f} INR Cr\n"
            f"  Written:    {written_cr:,.2f} INR Cr\n"
            f"  Gap:        {diff_cr:,.2f} INR Cr\n"
            f"  Fix category mapping then re-run. Downstream sheets NOT updated."
        )
        if raise_error:
            raise PipelineValidationError(msg)
        sys.exit(f"\n{'='*64}\n  {msg}\n{'='*64}")
    print("  [VALIDATE] ✓ GMV totals match — proceeding to downstream sheets.\n")


def build_change_summary(new_month: str, raw_data_dict: Dict, threshold_pp: float) -> str:
    return "\n".join([
        "=" * 68,
        f"  MEESHO MONTHLY UPDATE — {new_month}",
        "=" * 68, "",
        "  MODULES (in order):",
        f"    1. raw_working_sheet.py     → 4-col block ({len(raw_data_dict)} categories)",
        f"    2. raw_data.py              → data + MoM + YoY cols",
        f"    3. meesho_prelim.py         → Expert + RSC + YoY sections",
        f"    4. error_margin.py          → drag section end-cols",
        f"    5. fk_prelim.py             → drag section end-cols",
        f"    6. client_prelim.py         → copy prior month block +6 cols",
        f"    7. new_meesho_masterfile.py → contrib insert + data/MoM drag + hardcoded gen",
        "",
        "  OUTPUT:       Meesho may output.xlsx",
        "=" * 68,
    ])


def run_full_pipeline(
    wb,
    new_month: str,
    raw_data_dict: Dict[str, Tuple[int, int]],
    threshold_pp: float,
    dry_run: bool,
    save_step=None,
    stop_after: int = 7,
    *,
    raise_on_validation_error: bool = False,
    src_path: Optional[str] = None,
    fk_manual: Optional[FKManualInputs] = None,
    visible_months: Optional[Dict[str, Set[str]]] = None,
    visible_month_labels: Optional[Set[str]] = None,
    interactive: bool = True,
):
    """
    Execute steps 1–7 (optionally stopping earlier), passing anchor contracts forward.
    stop_after: halt after this step number (1–7).
    Returns (rws_result, rd_contract, mpv_contract, unmapped).
    """
    print("\n─── STEP 1: raw_working_sheet ───────────────────────")
    rws_result = raw_working_sheet.run(wb, new_month, raw_data_dict, dry_run)
    if not dry_run:
        validate_gmv_totals(
            raw_data_dict, rws_result["matched_by_name"],
            raise_error=raise_on_validation_error,
        )
        if save_step:
            save_step("Step 1")
    if stop_after <= 1:
        return rws_result, None, None, rws_result["unmapped"]

    print("\n─── STEP 2: raw_data ───────────────────────────────")
    rd_contract = raw_data.run(wb, new_month, rws_result, threshold_pp, dry_run)
    if not dry_run and save_step:
        save_step("Step 2")
    if stop_after <= 2:
        return rws_result, rd_contract, None, rws_result["unmapped"]

    print("\n─── STEP 3: meesho_prelim ─────────────────────────")
    mpv_contract = meesho_prelim.run(wb, new_month, rd_contract, threshold_pp, dry_run)
    if not dry_run and save_step:
        save_step("Step 3")
    if stop_after <= 3:
        return rws_result, rd_contract, mpv_contract, rws_result["unmapped"]

    print("\n─── STEP 4: error_margin ──────────────────────────")
    em_contract = error_margin.run(wb, new_month, mpv_contract, dry_run, src_path=src_path)
    if not dry_run and save_step:
        save_step("Step 4")
    if stop_after <= 4:
        return rws_result, rd_contract, mpv_contract, rws_result["unmapped"]

    print("\n─── STEP 5: fk_prelim ─────────────────────────────")
    fk_contract = fk_prelim.run(
        wb, new_month, mpv_contract, dry_run,
        fk_manual=fk_manual, interactive=interactive,
    )
    if not dry_run and save_step:
        save_step("Step 5")
    if stop_after <= 5:
        return rws_result, rd_contract, mpv_contract, rws_result["unmapped"]

    print("\n─── STEP 6: client_prelim ─────────────────────────")
    client_prelim.run(wb, new_month, mpv_contract, dry_run)
    if not dry_run and save_step:
        save_step("Step 6")
    if stop_after <= 6:
        return rws_result, rd_contract, mpv_contract, rws_result["unmapped"]

    print("\n─── STEP 7: new_meesho_masterfile ─────────────────")
    new_meesho_masterfile.run(
        wb, new_month, mpv_contract, em_contract, fk_contract, dry_run,
        src_path=src_path,
    )
    if not dry_run and save_step:
        save_step("Step 7")

    if not dry_run and stop_after >= 7:
        hide_month_columns_interactive(
            wb,
            visible_months=visible_months,
            blanket_labels=visible_month_labels,
            interactive=interactive,
        )

    return rws_result, rd_contract, mpv_contract, rws_result["unmapped"]


def execute_pipeline(
    master_path: str,
    raw_data_path: str,
    *,
    new_month: str,
    threshold_pp: float = 45.0,
    raw_sheet: Optional[str] = None,
    total_gmv_check: Optional[float] = None,
    stage: PipelineStage = "client_prelim",
    output_path: Optional[str] = None,
    fk_manual: Optional[FKManualInputs] = None,
    visible_months: Optional[Dict[str, Set[str]]] = None,
    visible_month_labels: Optional[Set[str]] = None,
) -> Tuple[str, list, Dict]:
    """
    Programmatic entry point for CLI and web UI.
    Returns (output_path, warnings, info_dict).
    """
    if not os.path.exists(master_path):
        raise FileNotFoundError(f"Workbook not found: {master_path}")
    if not os.path.exists(raw_data_path):
        raise FileNotFoundError(f"Raw data file not found: {raw_data_path}")

    raw_data_dict = load_raw_data(raw_data_path, sheet_name=raw_sheet)
    if not raw_data_dict:
        raise ValueError("No categories found in raw data file.")

    stop_after = STAGE_STOP_AFTER[stage]
    warnings: list = []

    if total_gmv_check is not None:
        raw_cr = sum(v[1] for v in raw_data_dict.values()) / 1e7
        if abs(raw_cr - total_gmv_check) > 0.05:
            warnings.append(
                f"Raw input GMV ({raw_cr:,.2f} Cr) differs from cross-check "
                f"({total_gmv_check:,.2f} Cr)."
            )

    out_dir = os.path.dirname(os.path.abspath(master_path)) or "."
    if output_path is None:
        suffix = {
            "raw_data": "through_raw_data",
            "error_margin": "through_error_margin",
            "client_prelim": "through_client_prelim",
            "full_master": "full_master",
        }[stage]
        output_path = os.path.join(
            out_dir,
            FINAL_OUTPUT_FILENAME if stage == "full_master" else f"Meesho {suffix} output.xlsx",
        )

    tmp_path = os.path.join(tempfile.gettempdir(), f"meesho_pipeline_{os.getpid()}.xlsx")
    shutil.copy2(master_path, tmp_path)
    wb = load_workbook(tmp_path)
    # Ensure Excel recalculates all formulas on open instead of showing stale values.
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"

    try:
        rws_result, _, _, unmapped = run_full_pipeline(
            wb, new_month, raw_data_dict, threshold_pp,
            dry_run=False, stop_after=stop_after,
            raise_on_validation_error=True,
            src_path=master_path,
            fk_manual=fk_manual,
            visible_months=visible_months,
            visible_month_labels=visible_month_labels,
            interactive=False,
        )
    except MonthSequenceError as exc:
        raise PipelineValidationError(str(exc)) from exc

    if total_gmv_check is not None and rws_result:
        written_cr = sum(v[1] for v in rws_result["matched_by_name"].values()) / 1e7
        if abs(written_cr - total_gmv_check) > 0.05:
            warnings.append(
                f"Written RWS GMV ({written_cr:,.2f} Cr) differs from cross-check "
                f"({total_gmv_check:,.2f} Cr)."
            )

    if unmapped:
        warnings.append(f"{len(unmapped)} unmapped categories — see orange cells in RWS.")

    wb.save(tmp_path)
    shutil.copy2(tmp_path, output_path)

    info = {
        "categories": len(raw_data_dict),
        "stage": stage,
        "stop_after": stop_after,
        "new_month": new_month,
        "unmapped": unmapped,
    }
    return output_path, warnings, info


def main():
    parser = argparse.ArgumentParser(description="Meesho monthly Excel update pipeline")
    parser.add_argument("workbook", nargs="?",
                        default="MeeshoMasterfilesampleautomation.xlsx")
    parser.add_argument("--month", required=True, help='e.g. "May\'26"')
    parser.add_argument("--raw-data", dest="raw_data_path",
                        help="Path to raw data (.xlsx or .csv)")
    parser.add_argument("--raw-sheet", dest="raw_sheet", default=None)
    parser.add_argument("--threshold", type=float, default=45.0,
                        help="MoM/YoY %% change threshold for red flag + note (default 45)")
    parser.add_argument("--output", dest="output_path", default=None,
                        help=f"Output file path (default: {FINAL_OUTPUT_FILENAME})")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--fk-asp", type=float, default=None,
                        help="FK Prelim ASP for new month (row 7); prompts if omitted")
    parser.add_argument("--fk-aov", type=float, default=None,
                        help="FK Prelim AoV for new month (row 8)")
    parser.add_argument("--fk-cancel", type=float, default=None,
                        help="FK Prelim Cancellation %% for new month (row 11); decimal or %%")
    parser.add_argument(
        "--visible-months",
        dest="visible_months_raw",
        default=None,
        help='Comma-separated month labels to keep visible (default: Apr\'25,May\'25,Apr\'26,May\'26)',
    )
    args = parser.parse_args()

    workbook_path = args.workbook
    new_month     = args.month
    threshold_pp  = args.threshold
    dry_run       = args.dry_run

    if not os.path.exists(workbook_path):
        sys.exit(f"ERROR: Workbook not found: {workbook_path}")
    if not args.raw_data_path:
        sys.exit("ERROR: No raw data file. Pass --raw-data your_file.csv")
    if not os.path.exists(args.raw_data_path):
        sys.exit(f"ERROR: Raw data file not found: {args.raw_data_path}")

    raw_data_dict = load_raw_data(args.raw_data_path, sheet_name=args.raw_sheet)
    print(f"Loaded {len(raw_data_dict)} categories from {args.raw_data_path}")
    if not raw_data_dict:
        sys.exit("ERROR: No raw data provided.")

    total_gmv_check = None
    if not dry_run:
        try:
            ans = input(
                f"\nEnter TOTAL GMV (INR Cr) for {new_month} cross-check "
                "(press Enter to skip): "
            ).strip()
            if ans:
                total_gmv_check = float(ans.replace(",", ""))
        except (ValueError, EOFError):
            pass

    fk_manual_cli: Optional[FKManualInputs] = None
    if args.fk_asp is not None and args.fk_aov is not None and args.fk_cancel is not None:
        from shared.fk_manual_inputs import parse_fk_manual_inputs
        fk_manual_cli = parse_fk_manual_inputs(
            str(args.fk_asp), str(args.fk_aov), str(args.fk_cancel),
        )

    visible_labels_cli = parse_visible_month_labels(args.visible_months_raw or "")

    print("\n" + build_change_summary(new_month, raw_data_dict, threshold_pp))

    if dry_run:
        print("\n[DRY-RUN MODE] Running partial analysis …\n")
        wb = load_workbook_safe(workbook_path, data_only=True)
        try:
            run_full_pipeline(wb, new_month, raw_data_dict, threshold_pp, dry_run=True)
        except MonthSequenceError as exc:
            sys.exit(f"\n  {exc}\n")
        print("\nDry-run complete. No files written.")
        return

    out_dir  = os.path.dirname(os.path.abspath(workbook_path)) or "."
    out_path = args.output_path if args.output_path else os.path.join(
        out_dir,
        FINAL_OUTPUT_FILENAME,
    )

    for path, label in [(out_path, "output file"), (workbook_path, "master file")]:
        if os.path.exists(path) or path == workbook_path:
            try:
                with open(path, "a+b" if os.path.exists(path) else "rb"):
                    pass
            except PermissionError:
                sys.exit(f"\nERROR: '{path}' ({label}) is open in Excel. Close it first.")

    print(f"\nOutput file: {out_path}")
    print("NOTE: Overwrites that file each run. Starts fresh from master.")
    ans = input("\nProceed? [yes/no]: ").strip().lower()
    if ans not in ("yes", "y"):
        print("Aborted.")
        return

    tmp_path = os.path.join(tempfile.gettempdir(), "meesho_monthly_work.xlsx")
    shutil.copy2(workbook_path, tmp_path)
    print("\nLoading workbook …")
    wb = load_workbook(tmp_path)
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.calcMode = "auto"

    def save_step(name: str) -> None:
        print(f"  Checkpoint after {name} …", end=" ", flush=True)
        wb.save(tmp_path)
        print("OK")

    try:
        _, _, _, unmapped = run_full_pipeline(
            wb, new_month, raw_data_dict, threshold_pp, dry_run=False,
            save_step=save_step,
            src_path=workbook_path,
            fk_manual=fk_manual_cli,
            visible_month_labels=visible_labels_cli,
            interactive=(fk_manual_cli is None),
        )
    except MonthSequenceError as exc:
        sys.exit(f"\n  {exc}\n")

    if unmapped:
        print(f"\n⚠  UNMAPPED ({len(unmapped)}):")
        for u in unmapped:
            print(f"   • {u}")

    print("\nFinal save …", end=" ", flush=True)
    wb.save(tmp_path)
    print("OK")

    try:
        with open(out_path, "a+b"):
            pass
    except PermissionError:
        sys.exit(
            f"\nERROR: '{out_path}' opened in Excel during run.\n"
            f"Work saved in: {tmp_path}"
        )

    shutil.copy2(tmp_path, out_path)
    next_month = next_month_label_after(new_month)
    print(f"\n✓ Done. Output: {out_path}")
    if next_month:
        print(
            f"\nNext month: use this output file as your master and run with "
            f"--month \"{next_month}\"."
        )
    print(
        "\nNext steps:\n"
        "  1. Open output in Excel → Ctrl+Alt+F9 to recalculate\n"
        "  2. Review YELLOW cells (trendline suggestions)\n"
        "  3. Review ORANGE cells (unmapped categories)\n"
        "  4. Run MakeCheckerZero macro (Error Margin checkers = 0)\n"
    )


if __name__ == "__main__":
    main()
