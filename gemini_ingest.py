"""
Gemini file-ingestion — convert messy uploads to Category/Orders/GMV CSV.

API key is passed per request only; never logged or persisted.
"""
import csv
import io
import mimetypes
import os
import re
from typing import Optional

GEMINI_CONVERT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".pdf", ".xlsx", ".xlsm", ".xls"}

PROMPT = """You are extracting Meesho monthly category sales data from the attached file.

Return ONLY a CSV with exactly these columns (header row required):
Category,Orders,GMV

Rules:
- Category: product category name as plain text
- Orders: integer count (no commas in values)
- GMV: integer in INR absolute units (not Crores)
- Include every category row you can identify; omit totals/subtotals rows
- No markdown, no code fences, no explanation — CSV text only
"""


def _strip_code_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```[a-zA-Z]*\n?", "", text)
        text = re.sub(r"\n?```$", "", text)
    return text.strip()


def _excel_to_text(path: str, max_rows: int = 200) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    lines = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            break
        cells = ["" if v is None else str(v) for v in row]
        lines.append(",".join(cells))
    wb.close()
    return "\n".join(lines)


def _normalize_csv(text: str) -> str:
    text = _strip_code_fences(text)
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        raise ValueError("Gemini returned empty CSV.")

    header = [c.strip().lower() for c in rows[0]]
    if header[:3] != ["category", "orders", "gmv"]:
        # Try to find header row
        start = 0
        for i, row in enumerate(rows):
            cols = [str(c).strip().lower() for c in row[:3]]
            if cols == ["category", "orders", "gmv"]:
                start = i
                break
        rows = rows[start:]
        if not rows:
            raise ValueError("Could not find Category,Orders,GMV header in Gemini output.")

    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(["Category", "Orders", "GMV"])
    for row in rows[1:]:
        if len(row) < 3:
            continue
        cat = str(row[0]).strip()
        if not cat or cat.lower() in ("category", "total", "total gmv"):
            continue
        writer.writerow([cat, row[1], row[2]])
    result = out.getvalue().strip()
    if result.count("\n") < 1:
        raise ValueError("Gemini CSV has no data rows.")
    return result + "\n"


def convert_to_raw_csv(api_key: str, file_path: str) -> str:
    """
    Send file to Gemini and return normalized CSV string (Category,Orders,GMV).
    api_key is used in-memory only for this call.
    """
    if not api_key or not api_key.strip():
        raise ValueError("Google API key is required.")

    ext = os.path.splitext(file_path)[1].lower()
    if ext not in GEMINI_CONVERT_EXTENSIONS:
        raise ValueError(
            f"Gemini convert supports: {', '.join(sorted(GEMINI_CONVERT_EXTENSIONS))}"
        )

    try:
        import google.generativeai as genai
    except ImportError as exc:
        raise ImportError(
            "Install google-generativeai for Gemini conversion: "
            "pip install google-generativeai"
        ) from exc

    genai.configure(api_key=api_key.strip())
    model = genai.GenerativeModel("gemini-1.5-flash")

    if ext in (".xlsx", ".xlsm", ".xls"):
        sheet_text = _excel_to_text(file_path)
        response = model.generate_content(
            [PROMPT, f"Spreadsheet contents:\n\n{sheet_text}"]
        )
    else:
        mime, _ = mimetypes.guess_type(file_path)
        if not mime:
            mime = "application/octet-stream"
        uploaded = genai.upload_file(file_path, mime_type=mime)
        response = model.generate_content([PROMPT, uploaded])

    if not response or not response.text:
        raise ValueError("Gemini returned no text. Try a clearer file or different model.")

    return _normalize_csv(response.text)


def save_converted_csv(csv_text: str, dest_path: str) -> str:
    with open(dest_path, "w", newline="", encoding="utf-8") as f:
        f.write(csv_text)
    return dest_path
