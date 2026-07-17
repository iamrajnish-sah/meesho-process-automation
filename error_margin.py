"""
STEP 4 — Error Margin - Expert vs Report

Drags all parallel section end-columns forward by one month.
Returns EMContract listing new columns written.

After the drag, runs the EM gap analysis automatically:
  - Column references come DIRECTLY from the pipeline contract (section_ends /
    new_cols) — no re-scan of the sheet.  The "current month" label is always
    new_month, which is passed in from run_pipeline.py and originates from the
    user's input; it is never hardcoded here.
  - Numeric values are read from src_path (the master workbook file, before
    pipeline changes) with data_only=True so cached formula results are available.
  - ALL output is written to the far-right of the sheet only
    (cols new_cols[-1]+1 … new_cols[-1]+7).  Nothing is written adjacent to the
    RSC block: the RSC block grows one column right every month, and placing any
    output next to it guarantees a collision next cycle (BUG 2 fix).
  - Headers are labelled with new_month ("May'26", "Jun'26", …) so they always
    reflect the pipeline month that produced them.

Cancellation block (detected dynamically via row-4 "Cancellation" label):
  - Formula rows are left as dragged (including row 38 Raw-Data cross-check).
  - 14 hardcoded category cancel-% rows get new-month values generated from
    Apr→May Raw Data GMV direction, capped by each row's own historical MoM
    swing (and a hard 2pp ceiling).

Meesho RSC Data block (pipeline contract new_cols[1] / section_ends[1]):
  - Formula rows are left as dragged.
  - Hardcoded newest-month RSC cells get values generated within ±5% of the
    corresponding Expert value, prioritising historical RSC trend then Expert.
"""
import os
import random
import re
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import column_index_from_string as CN, get_column_letter as CL

from raw_data import RD_GMV_ROLLUP_ROWS
from raw_working_sheet import (
    RWS_CAT_FIRST_ROW, RWS_CAT_LAST_ROW,
    build_canonical_name_map, resolve_canonical_name,
)
from shared.anchors import SHEET_ANCHORS, find_all_month_cols_in_row
from shared.month_utils import anchor_search_token, year_ago_label_for, parse_month_label
from shared.contracts import EMContract, MPVContract
from shared.formula_utils import drag_column_forward, next_col, col_num, col_str
from shared.style_utils import copy_cell_style, YELLOW_FILL, ORANGE_FILL, add_comment
from shared.trend_hardcode import (
    generate_pp_bounded,
    write_hardcoded_cell,
)

SHEET_NAME        = "Error Margin - Expert vs Report"
HDR_ROW           = 5
EM_DATA_ROW_START = 5
EM_DATA_ROW_END   = 110
EM_CHECKER_ROW    = 4
EXPERT_START_COL  = 4    # column D (never changes)
CANCEL_LABEL_ROW  = 4
MPV_CANCEL_ROW    = 11   # Meesho Prelim View Cancellation % row

# ── Gap analysis constants ─────────────────────────────────────────────────
MOM_GAP_THRESHOLD = 6      # pp — flag if |Expert MoM% − RSC MoM%| > this
YOY_GAP_THRESHOLD = 6      # pp — same for YoY
EXPERT_WEIGHT     = 0.7    # weight in suggest_adjustment_placeholder
RSC_WEIGHT        = 1 - EXPERT_WEIGHT
MIN_GMV           = 5.0    # ignore rows whose values are below this

# ── Cancellation hardcoded generation ──────────────────────────────────────
CANCEL_HARD_CAP_PP = 2.0   # never move a cancel-% more than this (either way)
CANCEL_HIST_MONTHS = 6     # lookback for realistic MoM swing range

# EM cancel row → (label, Raw Data GMV rows to sum).
# Matches Expert-block formulas, e.g. S17 = 'Raw Data'!R47*(1-BB17).
CANCEL_HARDCODED: Dict[int, Tuple[str, Tuple[int, ...]]] = {
    17: ("MLE", (47,)),
    20: ("Men Fashion", (34,)),
    22: ("Ethnic Wear", (28, 29, 30)),
    23: ("Western wear", (31, 32, 33)),
    24: ("Kids", (40,)),
    26: ("Footwear", (37,)),
    28: ("Watch", (41,)),
    29: ("Wearables (Eyewear + BW)", (36,)),
    30: ("Luggage and travel Accessories & Handbags", (39,)),
    31: ("Jewellery", (38,)),
    33: ("Home Décor & Furnishings", (44,)),
    34: ("Kitchen, Home Improvement & Appliances", (45,)),
    35: ("Beauty&PersonalCare", (42,)),
    36: ("BGM", (50, 46)),
}

# Formula rows in the Cancellation block — dragged only, never overwritten.
CANCEL_FORMULA_ROWS = (15, 16, 18, 19, 21, 27, 32, 38)

# ── RSC hardcoded generation ───────────────────────────────────────────────
RSC_EXPERT_BOUND_PCT = 5.0   # mandatory ±5% of Expert (never violated)
RSC_HIST_MONTHS      = 6
RSC_TREND_W_RSC      = 0.55  # weight: historical RSC MoM trend
RSC_TREND_W_RATIO    = 0.30  # weight: historical RSC/Expert ratio
RSC_TREND_W_EXP      = 0.15  # weight: historical Expert MoM trend

# ── RSC-block assumption / contribution hardcoded rows (±1pp MoM cap) ────────
ASSUMPTION_HIST_MONTHS = 6
ASSUMPTION_PP_CAP      = 1.0   # 1 percentage point (native units via trend_hardcode)

EM_ASSUMPTION_FULL_ROWS    = (44, 65, 66, 67, 68, 69, 70)
EM_ASSUMPTION_NEWEST_ROWS  = (48, 52)
EM_CONTRIBUTION_ROWS       = tuple(range(85, 93))

MONTH_RE  = re.compile(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'\d{2}", re.I)
RED_FILL  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREEN_FILL= PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
GREY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BOLD      = Font(bold=True)


# ─────────────────────────────────────────────────────────────────────────────
# Pipeline entry point
# ─────────────────────────────────────────────────────────────────────────────

def run(wb, new_month: str, mpv_contract: MPVContract, dry_run: bool,
        src_path: Optional[str] = None) -> EMContract:
    """Step 4.

    src_path — path to the tmp/master workbook (used read-only for cached
    Expert/RSC numeric values during gap analysis only).
    """
    ws = wb[SHEET_NAME]
    anchor = SHEET_ANCHORS[SHEET_NAME]
    print(f"\n  [EM] Processing Error Margin sheet …")

    search = anchor_search_token(new_month)
    section_ends = find_all_month_cols_in_row(ws, anchor["header_row"], search)
    if not section_ends:
        print(f"  [EM] WARNING: No sections ending with {search!r}. Skipping.")
        return EMContract(section_new_cols=[])

    print(f"  [EM] Found {len(section_ends)} section(s): {section_ends}")
    new_cols: list = []
    for last_col in section_ends:
        new_col = next_col(last_col)
        new_cols.append(new_col)
        if dry_run:
            print(f"  [DRY-RUN] EM section: {last_col}→{new_col}")
            continue
        ws[f"{new_col}{anchor['header_row']}"].value = new_month
        copy_cell_style(ws[f"{last_col}{anchor['header_row']}"],
                        ws[f"{new_col}{anchor['header_row']}"])
        n = drag_column_forward(ws, last_col, new_col,
                                EM_DATA_ROW_START, EM_DATA_ROW_END, dry_run=dry_run)
        print(f"  [EM] ✓ Section {last_col}→{new_col}: {n} cells.")

    if not dry_run:
        print(f"  [EM] ⚠  Confirm row {EM_CHECKER_ROW} checkers = 0 in Excel.")
        try:
            _fill_cancellation_hardcoded(
                wb, ws, new_month, section_ends, new_cols, mpv_contract,
                src_path=src_path,
            )
        except Exception as exc:
            print(f"  [EM Cancel] ⚠ Cancellation fill skipped: {exc}")
        try:
            _fill_rsc_hardcoded(
                wb, ws, new_month, section_ends, new_cols, src_path=src_path,
            )
        except Exception as exc:
            print(f"  [EM RSC] ⚠ RSC hardcoded fill skipped: {exc}")
        try:
            _fill_assumption_hardcoded(
                ws, new_month, section_ends, new_cols,
            )
        except Exception as exc:
            print(f"  [EM Assumption] ⚠ Assumption fill skipped: {exc}")

    if not dry_run and src_path and os.path.exists(src_path) and len(new_cols) >= 2:
        try:
            _run_gap_analysis(ws, new_month, section_ends, new_cols, src_path)
        except Exception as exc:
            print(f"  [EM Gap] ⚠ Gap analysis skipped: {exc}")

    return EMContract(section_new_cols=new_cols)


# ─────────────────────────────────────────────────────────────────────────────
# Cancellation block — hardcoded May values
# ─────────────────────────────────────────────────────────────────────────────

def _find_cancellation_cols(
    ws, section_ends: List[str], new_cols: List[str],
) -> Tuple[Optional[str], Optional[str]]:
    """Locate Cancellation prev/new columns dynamically (never hardcode AM/BB).

    Finds the 'Cancellation' label in row 4, walks contiguous month headers in
    row 5 to the section end, and matches that end against section_ends.
    """
    start_n = None
    for c in range(1, 200):
        v = ws.cell(CANCEL_LABEL_ROW, c).value
        if isinstance(v, str) and "ancel" in v.lower():
            start_n = c
            break
    if start_n is None:
        print("  [EM Cancel] No 'Cancellation' label in row 4 — skipping.")
        return None, None

    # After the section drag the new month header already sits in the next
    # column, so walk the contiguous block and pick the rightmost letter that
    # still appears in section_ends (the pre-drag end), not the newly written col.
    block_cols: List[str] = []
    for c in range(start_n, 220):
        v = ws.cell(HDR_ROW, c).value
        if v is None or (isinstance(v, str) and not str(v).strip()):
            break
        block_cols.append(CL(c))

    prev_col = None
    for letter in reversed(block_cols):
        if letter in section_ends:
            prev_col = letter
            break
    if prev_col is None:
        print(f"  [EM Cancel] No section_ends match in block {block_cols} — skipping.")
        return None, None

    idx = section_ends.index(prev_col)
    new_col = new_cols[idx]
    print(f"  [EM Cancel] Block detected: start={CL(start_n)}, "
          f"cols={block_cols[0]}…{block_cols[-1]}, prev={prev_col} → new={new_col}")
    return prev_col, new_col


def _rd_gmv_cols(ws, expert_prev_col: str, mpv_contract: MPVContract) -> Tuple[str, str]:
    """Resolve Raw Data prev/new GMV columns from Expert formula + MPV col_map."""
    sample = ws[f"{expert_prev_col}17"].value
    m = re.search(r"'Raw Data'!\$?([A-Z]+)\$?47", str(sample or ""))
    if not m:
        raise RuntimeError(
            f"Could not parse Raw Data col from {expert_prev_col}17={sample!r}"
        )
    rd_prev = m.group(1)
    rd_new = mpv_contract["col_map"].get(rd_prev) or next_col(rd_prev)
    return rd_prev, rd_new


def _hist_mom_swings_pp(ws, row: int, prev_col: str) -> List[float]:
    """MoM swings (percentage points) over the last CANCEL_HIST_MONTHS months."""
    prev_n = CN(prev_col)
    start_n = prev_n - (CANCEL_HIST_MONTHS - 1)
    vals: List[Optional[float]] = []
    for c in range(start_n, prev_n + 1):
        v = ws.cell(row, c).value
        try:
            vals.append(float(v))
        except (TypeError, ValueError):
            vals.append(None)
    deltas: List[float] = []
    for i in range(1, len(vals)):
        if vals[i] is not None and vals[i - 1] is not None:
            deltas.append((vals[i] - vals[i - 1]) * 100.0)
    return deltas


def _fill_cancellation_hardcoded(
    wb,
    ws,
    new_month: str,
    section_ends: List[str],
    new_cols: List[str],
    mpv_contract: MPVContract,
    src_path: Optional[str] = None,
) -> None:
    """Overwrite the 14 hardcoded cancel-% cells in the new Cancellation column.

    Formula rows (incl. row 38 Raw-Data pattern) stay as dragged.
    Row 16 Total is re-pointed to MPV Expert cancel% for the new month
    (AK→AL shift would miss MPV's actual Expert cancel cell).

    Apr GMV is read from the original master with data_only=True (cached Excel
    values).  The in-workbook Raw Data prev-month XLOOKUPs currently point at a
    stale RWS block, so live formula evaluation cannot be trusted for Apr.
    May GMV is resolved live from the post-pipeline Raw Data / RWS numbers.
    """
    prev_col, new_col = _find_cancellation_cols(ws, section_ends, new_cols)
    if not prev_col or not new_col:
        return

    expert_prev = section_ends[0]
    expert_new = new_cols[0]
    rd_prev, rd_new = _rd_gmv_cols(ws, expert_prev, mpv_contract)
    rd_prev_n, rd_new_n = CN(rd_prev), CN(rd_new)
    mpv_expert = mpv_contract["expert_col"]

    ws_rd_apr = None
    if src_path and os.path.exists(src_path):
        try:
            ws_rd_apr = openpyxl.load_workbook(src_path, data_only=True)["Raw Data"]
            print(f"  [EM Cancel] Apr GMV source: cached values from {os.path.basename(src_path)}")
        except Exception as exc:
            print(f"  [EM Cancel] WARNING: could not load Apr cached GMV ({exc})")

    def _gmv_may(rows: Tuple[int, ...]) -> Optional[float]:
        total = 0.0
        for r in rows:
            if r == 50:
                # Prefer Apr-compatible Others rollup SUM(51:56,60) — the May
                # formula SUM(51:60) double-counts 57:59 via nested S60.
                parts = [
                    _resolve_raw_data_gmv(wb, rr, rd_new_n)
                    for rr in (51, 52, 53, 54, 55, 56, 60)
                ]
                if any(p is None for p in parts):
                    return None
                total += sum(parts)  # type: ignore[arg-type]
                continue
            v = _resolve_raw_data_gmv(wb, r, rd_new_n)
            if v is None:
                return None
            total += v
        return total

    def _gmv_apr(rows: Tuple[int, ...]) -> Optional[float]:
        if ws_rd_apr is not None:
            total = 0.0
            for r in rows:
                v = ws_rd_apr.cell(r, rd_prev_n).value
                try:
                    total += float(v)
                except (TypeError, ValueError):
                    return None
            return total
        # Fallback — may be wrong if Raw Data prev XLOOKUP is stale
        total = 0.0
        for r in rows:
            v = _resolve_raw_data_gmv(wb, r, rd_prev_n)
            if v is None:
                return None
            total += v
        return total

    # Row 16 Total: point at MPV Expert cancel% for the new month (not AL11).
    # Apr pattern was BB16 → MPV!AK11 → T11; May must be MPV!{expert_col}11.
    total_formula = f"='Meesho Prelim View'!{mpv_expert}{MPV_CANCEL_ROW}"
    ws[f"{new_col}16"].value = total_formula
    copy_cell_style(ws[f"{prev_col}16"], ws[f"{new_col}16"])
    print(f"  [EM Cancel] Row 16 Total → {total_formula}")

    # Confirm row 38 kept the Raw-Data pattern (do not rewrite history).
    r38 = ws[f"{new_col}38"].value
    print(f"  [EM Cancel] Row 38 (May, Raw-Data pattern, untouched logic): {r38!r}")
    print(f"  [EM Cancel] Row 38 prior month left as-is: "
          f"{prev_col}38={ws[f'{prev_col}38'].value!r}")
    print(f"  [EM Cancel] Raw Data GMV cols: {rd_prev} (prev) → {rd_new} (new)")

    print(f"\n  [EM Cancel] Historical MoM swings (last {CANCEL_HIST_MONTHS}m, pp) "
          f"+ Apr→May GMV → generated {new_month} cancel%:")
    print(f"  {'Row':>4} {'Label':40} {'Apr%':>8} {'GMVΔ%':>8} {'dir':>4} "
          f"{'histMax':>8} {'Δpp':>7} {'May%':>8}  hist_deltas")

    rng = random.Random(f"cancel|{new_month}")
    generated: Dict[int, float] = {}
    leverage: List[Tuple[float, int, str, float]] = []

    for row, (label, rd_rows) in CANCEL_HARDCODED.items():
        apr_pct = ws[f"{prev_col}{row}"].value
        try:
            apr_pct_f = float(apr_pct)
        except (TypeError, ValueError):
            print(f"  [EM Cancel] WARNING: {prev_col}{row} not numeric ({apr_pct!r}) — skip")
            continue

        gmv_apr = _gmv_apr(rd_rows)
        gmv_may = _gmv_may(rd_rows)
        if gmv_apr is None or gmv_may is None:
            print(f"  [EM Cancel] WARNING: could not resolve GMV for r{row} "
                  f"({label}) via Raw Data {rd_rows} "
                  f"(apr={gmv_apr}, may={gmv_may}) — skip")
            continue

        leverage.append((abs(gmv_may), row, label, gmv_may))

        if gmv_apr > 0:
            gmv_chg_pct = (gmv_may / gmv_apr - 1.0) * 100.0
        else:
            gmv_chg_pct = 0.0

        if gmv_may > gmv_apr * (1 + 1e-12):
            direction = +1
            dir_s = "UP"
        elif gmv_may < gmv_apr * (1 - 1e-12):
            direction = -1
            dir_s = "DN"
        else:
            direction = 0
            dir_s = "FLAT"

        deltas = _hist_mom_swings_pp(ws, row, prev_col)
        hist_max = max((abs(d) for d in deltas), default=0.3)
        cap_pp = min(CANCEL_HARD_CAP_PP, max(hist_max, 0.05))

        if direction == 0:
            delta_pp = 0.0
        else:
            # Magnitude only randomized; direction follows GMV.
            delta_pp = direction * rng.uniform(cap_pp * 0.25, cap_pp)

        may_pct = max(0.0, min(0.5, apr_pct_f + delta_pp / 100.0))
        generated[row] = may_pct

        cell = ws[f"{new_col}{row}"]
        cell.value = may_pct
        copy_cell_style(ws[f"{prev_col}{row}"], cell)
        cell.fill = YELLOW_FILL
        add_comment(
            ws, f"{new_col}{row}",
            f"[EM Cancel] Generated for {new_month}.\n"
            f"Apr={apr_pct_f:.4%} → May={may_pct:.4%} (Δ={delta_pp:+.3f}pp).\n"
            f"GMV {dir_s}: {gmv_apr:.2f} → {gmv_may:.2f} Cr ({gmv_chg_pct:+.2f}%).\n"
            f"Hist max |MoM|={hist_max:.3f}pp; cap={cap_pp:.3f}pp "
            f"(hard≤{CANCEL_HARD_CAP_PP}pp).",
        )

        d_str = ",".join(f"{d:+.2f}" for d in deltas) if deltas else "n/a"
        print(f"  {row:>4} {label:40} {apr_pct_f*100:7.3f}% {gmv_chg_pct:+7.2f}% "
              f"{dir_s:>4} {hist_max:7.3f} {delta_pp:+6.3f} {may_pct*100:7.3f}%  [{d_str}]")

    # ── Validation: row 16 vs row 38, Expert Total vs row 38 ──
    mpv_ws = wb["Meesho Prelim View"]
    total_cancel = mpv_ws[f"{mpv_expert}{MPV_CANCEL_ROW}"].value
    try:
        total_cancel_f = float(total_cancel)
    except (TypeError, ValueError):
        total_cancel_f = None

    # Total GMV = Fashion + BPC + Home + Grocery + Electronics + Others
    # (same parts as Raw Data row 25; Others via Apr-compatible 51:56+60)
    _tot_parts = [
        _resolve_raw_data_gmv(wb, r, rd_new_n) for r in (26, 42, 43, 46, 47)
    ]
    _others = _gmv_may((50,))
    if all(p is not None for p in _tot_parts) and _others is not None:
        total_gmv_may = sum(_tot_parts) + _others  # type: ignore[operator]
    else:
        total_gmv_may = sum(g for _, _, _, g in leverage) if leverage else 0.0

    row38_implied = (
        total_gmv_may * (1.0 - total_cancel_f)
        if total_cancel_f is not None else None
    )

    def shipped(row_cancel: int, rd_rows: Tuple[int, ...]) -> float:
        gmv = _gmv_may(rd_rows) or 0.0
        pct = generated.get(row_cancel)
        if pct is None:
            try:
                pct = float(ws[f"{new_col}{row_cancel}"].value)
            except (TypeError, ValueError):
                pct = 0.0
        return gmv * (1.0 - pct)

    mle = shipped(17, (47,))
    men = shipped(20, (34,))
    ethnic = shipped(22, (28, 29, 30))
    western = shipped(23, (31, 32, 33))
    kids = shipped(24, (40,))
    foot = shipped(26, (37,))
    watch = shipped(28, (41,))
    wear = shipped(29, (36,))
    lug = shipped(30, (39,))
    jew = shipped(31, (38,))
    home = shipped(33, (44,))
    kit = shipped(34, (45,))
    bpc = shipped(35, (42,))
    bgm = shipped(36, (50, 46))

    women = ethnic + western
    apparel = men + women + kids
    accessories = watch + wear + lug + jew
    fashion = apparel + foot + accessories
    home_kit = home + kit
    expert_total_approx = mle + fashion + home_kit + bpc + bgm

    print(f"\n  [EM Cancel] VALIDATION — {new_month}")
    if total_cancel_f is not None:
        print(f"  Row 16 Total cancel%  ({new_col}16): {total_cancel_f*100:.4f}%")
    else:
        print(f"  Row 16 Total cancel%  ({new_col}16): {total_cancel!r}")
    if row38_implied is not None:
        print(f"  Row 38 cross-check GMV ({new_col}38): {row38_implied:,.4f} Cr "
              f"[= {rd_new}25*(1-{new_col}16)]")
    else:
        print("  Row 38: could not compute")
    print(f"  Expert Total approx   ({expert_new}16): {expert_total_approx:,.4f} Cr "
          f"(from category cancel% × May GMV)")
    if row38_implied is not None and row38_implied != 0:
        gap = expert_total_approx - row38_implied
        gap_pct = gap / row38_implied * 100.0
        print(f"  Checker-style gap (Expert Total − row38): {gap:+,.4f} Cr "
              f"({gap_pct:+.3f}%)")
        if abs(gap_pct) > 1.0:
            print("  ⚠ DIVERGENCE >1%: row 16 comes from MPV trendline Total, "
                  "while Expert Total uses the 14 generated category rates. "
                  "They are not forced to reconcile.")

    print(f"\n  [EM Cancel] Checkers / forward deps on Cancellation block:")
    print(f"  • Row 59 '{ws['B59'].value}' — formula in prev month "
          f"{prev_col}59={ws[f'{prev_col}59'].value!r}")
    print(f"    After drag: {new_col}59={ws[f'{new_col}59'].value!r}  "
          f"(Expert Total − row38 cross-check; expect ~0 when Excel recalcs)")
    print(f"  • Expert leaf rows ({expert_new}17 etc.) reference "
          f"(1-{new_col}{{row}}) — written by section drag, not modified here.")
    print(f"  • Rows 109/135 Checker are RSC-block formulas "
          f"(V:AJ area) — do not reference Cancellation AM:{prev_col}.")

    leverage.sort(reverse=True)
    print(f"\n  [EM Cancel] Highest-GMV leverage categories "
          f"(for future target tuning — do NOT force unrealistic swings now):")
    for i, (_, row, label, gmv) in enumerate(leverage[:8], 1):
        share = (gmv / total_gmv_may * 100.0) if total_gmv_may else 0.0
        print(f"    {i}. r{row} {label:40} May GMV={gmv:8.2f} Cr  ({share:5.2f}% of Total)")

    print(f"  [EM Cancel] Formula rows left as dragged only: {CANCEL_FORMULA_ROWS}")
    print(f"  [EM Cancel] Hardcoded rows written: {sorted(generated)}")


# ─────────────────────────────────────────────────────────────────────────────
# Meesho RSC Data — hardcoded newest-month values
# ─────────────────────────────────────────────────────────────────────────────

def _hist_value_series(
    ws, row: int, end_col_n: int, months: int,
) -> List[Optional[float]]:
    """Return oldest→newest numeric values for the last `months` columns."""
    series: List[Optional[float]] = []
    for i in range(months - 1, -1, -1):
        col_n = end_col_n - i
        if col_n < 1:
            series.append(None)
        else:
            series.append(_val(ws, row, col_n))
    return series


def _growth_rates(vals: List[Optional[float]]) -> List[float]:
    rates: List[float] = []
    for i in range(1, len(vals)):
        a, b = vals[i - 1], vals[i]
        if a is not None and b is not None and a > 0:
            rates.append(b / a - 1.0)
    return rates


def _fill_rsc_hardcoded(
    wb,
    ws,
    new_month: str,
    section_ends: List[str],
    new_cols: List[str],
    src_path: Optional[str] = None,
) -> None:
    """Generate realistic values for hardcoded Meesho RSC cells in the new month.

    Only touches rows where the new RSC column is a bare number (not a formula).
    Hard constraint: every generated value stays within ±5% of Expert May.
    """
    if len(section_ends) < 2 or len(new_cols) < 2:
        print("  [EM RSC] Need Expert + RSC sections in contract — skipping.")
        return

    exp_prev_col = section_ends[0]
    exp_new_col  = new_cols[0]
    rsc_prev_col = section_ends[1]
    rsc_new_col  = new_cols[1]
    exp_prev_n   = CN(exp_prev_col)
    exp_new_n    = CN(exp_new_col)
    rsc_prev_n   = CN(rsc_prev_col)
    rsc_new_n    = CN(rsc_new_col)

    print(f"\n  [EM RSC] Meesho RSC block: {rsc_prev_col} → {rsc_new_col} "
          f"(Expert {exp_new_col} for ±5% bound)")

    ws_prev = None
    if src_path and os.path.exists(src_path):
        try:
            ws_prev = openpyxl.load_workbook(src_path, data_only=True)[SHEET_NAME]
        except Exception as exc:
            print(f"  [EM RSC] WARNING: cannot load prev cached values ({exc})")

    bound = RSC_EXPERT_BOUND_PCT / 100.0
    rng = random.Random(f"rsc|{new_month}")

    print(f"  [EM RSC] Generating hardcoded {new_month} values "
          f"(±{RSC_EXPERT_BOUND_PCT:.0f}% Expert band, {RSC_HIST_MONTHS}m history):")
    print(f"  {'Row':>4} {'Label':38} {'ExpMay':>10} {'RscApr':>10} "
          f"{'Trend':>10} {'MayRSC':>10} {'vsExp%':>8}")

    generated: Dict[int, float] = {}
    skipped_formula = 0
    skipped_no_expert = 0

    for r in range(EM_DATA_ROW_START, EM_DATA_ROW_END + 1):
        # Meesho RSC category table only — skip adjustment rows below ~40
        if r > 40:
            continue
        if not _is_bare(ws, r, rsc_new_n):
            if ws.cell(r, rsc_new_n).value is not None:
                skipped_formula += 1
            continue

        label = str(ws.cell(r, 2).value or ws.cell(r, 1).value or f"row{r}")[:38]

        expert_may = _val_current(ws, r, exp_new_n)
        if expert_may is None or expert_may <= 0:
            skipped_no_expert += 1
            print(f"  {r:>4} {label:38}  SKIP — Expert {exp_new_col} not resolved")
            continue

        expert_prev = _val(ws_prev, r, exp_prev_n) if ws_prev else None
        rsc_prev    = _val(ws_prev, r, rsc_prev_n) if ws_prev else _val(ws, r, rsc_prev_n)
        if expert_prev is None or expert_prev <= 0 or rsc_prev is None or rsc_prev <= 0:
            skipped_no_expert += 1
            print(f"  {r:>4} {label:38}  SKIP — missing prev Expert/RSC history")
            continue

        rsc_hist = (
            _hist_value_series(ws_prev, r, rsc_prev_n, RSC_HIST_MONTHS)
            if ws_prev else [rsc_prev]
        )
        exp_hist = (
            _hist_value_series(ws_prev, r, exp_prev_n, RSC_HIST_MONTHS)
            if ws_prev else [expert_prev]
        )

        avg_rsc_g = (
            sum(_growth_rates(rsc_hist)) / len(_growth_rates(rsc_hist))
            if _growth_rates(rsc_hist) else 0.0
        )
        avg_exp_g = (
            sum(_growth_rates(exp_hist)) / len(_growth_rates(exp_hist))
            if _growth_rates(exp_hist) else 0.0
        )

        ratios: List[float] = []
        for rv, ev in zip(rsc_hist, exp_hist):
            if rv is not None and ev is not None and ev > 0:
                ratios.append(rv / ev)
        avg_ratio = (
            sum(ratios) / len(ratios) if ratios else rsc_prev / expert_prev
        )

        # Priority 2 & 3: trend-based raw target (before ±5% clamp)
        rsc_trend    = rsc_prev * (1.0 + avg_rsc_g)
        ratio_trend  = expert_may * avg_ratio
        expert_trend = rsc_prev * (1.0 + avg_exp_g)
        raw_target = (
            RSC_TREND_W_RSC   * rsc_trend
            + RSC_TREND_W_RATIO * ratio_trend
            + RSC_TREND_W_EXP   * expert_trend
        )

        # Priority 1: mandatory ±5% Expert band
        lo = expert_may * (1.0 - bound)
        hi = expert_may * (1.0 + bound)
        clamped = max(lo, min(hi, raw_target))

        # Small magnitude-only noise inside the band (never breaks ±5%)
        band_w = hi - lo
        noise  = rng.uniform(-0.10, 0.10) * band_w
        final  = max(lo, min(hi, clamped + noise))

        cell = ws.cell(r, rsc_new_n)
        cell.value = round(final, 4)
        copy_cell_style(ws.cell(r, rsc_prev_n), cell)
        cell.fill = YELLOW_FILL
        pct_vs = (final - expert_may) / expert_may * 100.0
        add_comment(
            ws, f"{rsc_new_col}{r}",
            f"[EM RSC] Generated for {new_month}.\n"
            f"Expert {exp_new_col}={expert_may:.4f}; band "
            f"[{lo:.4f}, {hi:.4f}] (±{RSC_EXPERT_BOUND_PCT:.0f}%).\n"
            f"Trend raw={raw_target:.4f} "
            f"(RSC MoM avg={avg_rsc_g:+.2%}, ratio={avg_ratio:.4f}, "
            f"Exp MoM avg={avg_exp_g:+.2%}).\n"
            f"Final={final:.4f} ({pct_vs:+.2f}% vs Expert).",
        )
        generated[r] = final

        print(f"  {r:>4} {label:38} {expert_may:10.2f} {rsc_prev:10.2f} "
              f"{raw_target:10.2f} {final:10.2f} {pct_vs:+7.2f}%")

    # Validation — confirm ±5% never violated
    violations = []
    for r, rval in generated.items():
        ev = _val_current(ws, r, exp_new_n)
        if ev and ev > 0:
            pct = abs((rval - ev) / ev * 100.0)
            if pct > RSC_EXPERT_BOUND_PCT + 1e-6:
                violations.append((r, pct))

    print(f"\n  [EM RSC] Hardcoded rows written: {len(generated)} "
          f"(formula rows untouched: {skipped_formula})")
    if skipped_no_expert:
        print(f"  [EM RSC] Skipped (no Expert/history): {skipped_no_expert}")
    if violations:
        print(f"  [EM RSC] ⚠ BOUND VIOLATIONS: {violations}")
    else:
        print(f"  [EM RSC] ✓ All {len(generated)} values within "
              f"±{RSC_EXPERT_BOUND_PCT:.0f}% of Expert {exp_new_col}")


# ─────────────────────────────────────────────────────────────────────────────
# RSC-block assumption & contribution % — hardcoded newest-month values
# ─────────────────────────────────────────────────────────────────────────────

def _assumption_row_label(ws, row: int) -> str:
    return str(ws.cell(row, 3).value or ws.cell(row, 2).value or f"row{row}")[:38]


def _should_fill_assumption(ws, row: int, prev_col_n: int, mode: str) -> bool:
    if not _is_bare(ws, row, prev_col_n):
        return False
    if mode == "full":
        return True
    if mode == "newest":
        prev_prev = prev_col_n - 1
        if prev_prev < 1:
            return False
        return not _is_bare(ws, row, prev_prev)
    if mode == "contribution":
        return True
    return False


def _fill_assumption_hardcoded(
    ws,
    new_month: str,
    section_ends: List[str],
    new_cols: List[str],
) -> None:
    """Generate May hardcoded cells in the Meesho RSC column block only.

    Touches rows 44, 48, 52, 65–70, 85–92 — never formula cells.
    """
    if len(section_ends) < 2 or len(new_cols) < 2:
        print("  [EM Assumption] Need RSC section in contract — skipping.")
        return

    rsc_prev_col = section_ends[1]
    rsc_new_col  = new_cols[1]
    rsc_prev_n   = CN(rsc_prev_col)
    rsc_new_n    = CN(rsc_new_col)

    print(f"\n  [EM Assumption] RSC block {rsc_prev_col} → {rsc_new_col} "
          f"(±1pp MoM cap, {ASSUMPTION_HIST_MONTHS}m trend):")
    print(f"  {'Row':>4} {'Label':38} {'Prev':>12} {'May':>12} "
          f"{'Δpp':>8} {'Trend':>5} {'OK':>4}")

    rng = random.Random(f"assumption|{new_month}")
    results: List[Dict] = []

    row_modes = (
        [(r, "full") for r in EM_ASSUMPTION_FULL_ROWS]
        + [(r, "newest") for r in EM_ASSUMPTION_NEWEST_ROWS]
        + [(r, "contribution") for r in EM_CONTRIBUTION_ROWS]
    )

    for row, mode in row_modes:
        if not _should_fill_assumption(ws, row, rsc_prev_n, mode):
            continue
        if not _is_bare(ws, row, rsc_new_n):
            continue

        new_val, meta = generate_pp_bounded(
            ws, row, rsc_prev_n,
            months=ASSUMPTION_HIST_MONTHS,
            rng=rng,
            floor=0.0 if mode == "contribution" else None,
        )
        if new_val is None or not meta:
            print(f"  {row:>4} {_assumption_row_label(ws, row):38}  SKIP — no prior value")
            continue

        write_hardcoded_cell(
            ws, row, rsc_new_n, new_val,
            tag="EM Assumption",
            prev_col=rsc_prev_col,
            new_col=rsc_new_col,
            meta=meta,
        )
        ok = "yes" if meta["followed_trend"] else "no"
        print(
            f"  {row:>4} {_assumption_row_label(ws, row):38} "
            f"{meta['prev']:12.6g} {new_val:12.6g} "
            f"{meta['pp_change']:+8.3f} {meta['trend_dir']:>5} {ok:>4}"
        )
        results.append({"row": row, **meta})

    print(f"  [EM Assumption] Hardcoded rows written: {len(results)} "
          f"(formula cells untouched)")


# ─────────────────────────────────────────────────────────────────────────────
# Gap analysis — helpers
# ─────────────────────────────────────────────────────────────────────────────

def _val(ws, row: int, col: int) -> Optional[float]:
    v = ws.cell(row, col).value
    try:
        f = float(v)
        return None if f != f else f
    except (TypeError, ValueError):
        return None


def _pct(new_val: float, base_val: float) -> Optional[float]:
    return None if base_val == 0 else (new_val - base_val) / abs(base_val) * 100


def _is_bare(ws_f, row: int, col: int) -> bool:
    return isinstance(ws_f.cell(row, col).value, (int, float))


def suggest_adjustment_placeholder(rsc_prev: float,
                                   expert_mom_rate: float,
                                   rsc_mom_rate:    float) -> float:
    """PLACEHOLDER — replace only this function body when wiring Gemini later.
    Signature must stay the same.

    Rule: RSC_prev × (1 + 0.7×Expert_MoM + 0.3×RSC_MoM)
    """
    return rsc_prev * (1 + EXPERT_WEIGHT * expert_mom_rate + RSC_WEIGHT * rsc_mom_rate)


_EXPERT_FORMULA_RE = re.compile(
    r"^='Raw Data'!([A-Z]+)(\d+)\*\(1-([A-Z]+)(\d+)\)$"
)
_CELL_REF_RE = re.compile(r"([A-Z]+)(\d+)")
_RD_SHEET_REF_RE = re.compile(r"'Raw Data'!([A-Z]+)(\d+)")


def _eval_expert_gmv_expr(wb, expr: str) -> Optional[float]:
    """Evaluate the GMV portion of an Expert shipped-GMV formula."""
    expr = expr.strip().strip("(").strip()
    if expr.upper().startswith("SUM("):
        inner = expr[4:].rstrip(")").strip()
        if ":" in inner and "'Raw Data'!" in inner:
            m = re.search(
                r"'Raw Data'!([A-Z]+)(\d+):([A-Z]+)(\d+)", inner, re.I,
            )
            if m:
                col_n = CN(m.group(1))
                r0, r1 = int(m.group(2)), int(m.group(4))
                total = 0.0
                for r in range(r0, r1 + 1):
                    v = _resolve_raw_data_gmv(wb, r, col_n)
                    if v is None:
                        return None
                    total += v
                return total
    refs = list(_RD_SHEET_REF_RE.finditer(expr))
    if refs:
        total = 0.0
        for m in refs:
            v = _resolve_raw_data_gmv(wb, int(m.group(2)), CN(m.group(1)))
            if v is None:
                return None
            total += v
        return total
    return None


def _rws_gmv_cr_value(wb, rws_row: int, gmv_cr_col: str) -> Optional[float]:
    """GMV (INR Cr) from a Raw Working Sheet row — handles BO = BM/10^7 style."""
    rws = wb["Raw Working Sheet"]
    c = CN(gmv_cr_col)
    v = _val(rws, rws_row, c)
    if v is not None:
        return v
    formula = rws.cell(rws_row, c).value
    if isinstance(formula, str) and formula.startswith("="):
        m = re.match(r"=([A-Z]+)(\d+)/10", formula.strip())
        if m:
            base = _val(rws, int(m.group(2)), CN(m.group(1)))
            return None if base is None else base / 1e7
    return None


def _resolve_raw_data_gmv(
    wb, row: int, col: int,
    cache: Optional[Dict[Tuple[int, int], Optional[float]]] = None,
    _stack: Optional[set] = None,
) -> Optional[float]:
    """Evaluate a Raw Data GMV cell (rollup, XLOOKUP, or bare number).

    Uses a value cache so overlapping SUM ranges (e.g. S50=SUM(S51:S60) where
    S60 itself is SUM(S57:S59)) reuse already-computed cells instead of
    returning None from a too-aggressive 'seen' set.
    """
    if cache is None:
        cache = {}
    if _stack is None:
        _stack = set()
    key = (row, col)
    if key in cache:
        return cache[key]
    if key in _stack:
        return None  # true circular reference
    _stack.add(key)

    rd = wb["Raw Data"]
    col_l = CL(col)
    result: Optional[float] = None

    v = _val(rd, row, col)
    if v is not None:
        result = v
    elif row in RD_GMV_ROLLUP_ROWS:
        result = _eval_rd_gmv_expr(
            wb, RD_GMV_ROLLUP_ROWS[row].replace("{{c}}", col_l), cache, _stack,
        )
    else:
        formula = rd.cell(row, col).value
        if isinstance(formula, str) and formula.startswith("="):
            body = formula[1:].strip()
            if "XLOOKUP" in body.upper():
                matches = re.findall(
                    r"'Raw Working Sheet'!\$([A-Z]+)\$5:\$\1\$34",
                    body, re.I,
                )
                gmv_cr_col = (
                    matches[1] if len(matches) >= 2
                    else (matches[0] if matches else "BO")
                )
                cat = rd.cell(row, 1).value
                if cat:
                    rws = wb["Raw Working Sheet"]
                    name_map = build_canonical_name_map(rws)
                    cat_s = str(cat).strip()
                    for r in range(RWS_CAT_FIRST_ROW, RWS_CAT_LAST_ROW + 1):
                        if resolve_canonical_name(rws, r, name_map) == cat_s:
                            result = _rws_gmv_cr_value(wb, r, gmv_cr_col)
                            break
            else:
                result = _eval_rd_gmv_expr(wb, body, cache, _stack)

    cache[key] = result
    _stack.discard(key)
    return result


def _eval_rd_gmv_expr(
    wb, expr: str,
    cache: Dict[Tuple[int, int], Optional[float]],
    _stack: set,
) -> Optional[float]:
    """Evaluate =S48+S49 or =SUM(S51:S60) style Raw Data expressions."""
    expr = expr.strip()
    if expr.upper().startswith("SUM(") and expr.endswith(")"):
        inner = expr[4:-1]
        if ":" in inner:
            m = _CELL_REF_RE.match(inner.split(":")[0].strip())
            m2 = _CELL_REF_RE.search(inner.split(":")[1])
            if not m or not m2:
                return None
            col_n = CN(m.group(1))
            r0, r1 = int(m.group(2)), int(m2.group(2))
            total = 0.0
            for r in range(r0, r1 + 1):
                v = _resolve_raw_data_gmv(wb, r, col_n, cache, _stack)
                if v is None:
                    return None
                total += v
            return total
    total = 0.0
    for m in _CELL_REF_RE.finditer(expr):
        v = _resolve_raw_data_gmv(wb, int(m.group(2)), CN(m.group(1)), cache, _stack)
        if v is None:
            return None
        total += v
    return total if _CELL_REF_RE.search(expr) else None


def _val_current(ws, row: int, col: int) -> Optional[float]:
    """Numeric value from the post-drag output sheet (newest month column).

    Bare numbers return directly. Expert formula cells (=Raw Data!…*(1-BC…))
    are evaluated from live sheet data because openpyxl has no cached value
    for formulas written in this same pipeline run.
    """
    v = _val(ws, row, col)
    if v is not None:
        return v
    raw = ws.cell(row, col).value
    if not isinstance(raw, str) or not raw.startswith("="):
        return None
    body = raw[1:].strip()

    m = _EXPERT_FORMULA_RE.match(body)
    if m:
        rd_col, rd_row = m.group(1), int(m.group(2))
        em_col, em_row = m.group(3), int(m.group(4))
        gmv  = _resolve_raw_data_gmv(ws.parent, rd_row, CN(rd_col))
        rate = _val(ws, em_row, CN(em_col))
        if gmv is None or rate is None:
            return None
        return gmv * (1 - rate)

    # SUM('Raw Data'!S28:S30)*(1-BC22) or (('Raw Data'!S50+'Raw Data'!S46)*(1-BC36))
    tail = re.search(r"\*\(1-(\$?)([A-Z]+)(\$?)(\d+)\)\)*$", body)
    if tail:
        em_col = CN(tail.group(2))
        em_row = int(tail.group(4))
        gmv_part = body[: tail.start()].strip().rstrip("(")
        gmv = _eval_expert_gmv_expr(ws.parent, gmv_part)
        rate = _val(ws, em_row, em_col)
        if gmv is not None and rate is not None:
            return gmv * (1 - rate)

    return None


def _find_category_rows(
    ws_curr, ws_f_curr,
    ws_prev, ws_f_prev,
    exp_n: int, exp_p: int,
    rsc_n: int, rsc_p: int,
    exp_yoy: Optional[int],
    rsc_yoy: Optional[int],
) -> List[Tuple]:
    """Rows where THIS month's RSC column (rsc_n on ws_curr) is a bare number.

    Current-month values (exp_n, rsc_n) come from ws_curr (post-drag output).
    Previous-month values (exp_p, rsc_p) come from ws_prev (master cached).
    YoY bases come from ws_prev (master)."""
    results = []
    for r in range(2, 150):
        b = ws_prev.cell(r, 2).value
        label = str(b).strip() if b else ""
        if "checker" in label.lower():
            continue
        ev  = _val_current(ws_curr, r, exp_n)
        epv = _val(ws_prev, r, exp_p)
        rv  = _val(ws_curr, r, rsc_n)
        rpv = _val(ws_prev, r, rsc_p)
        if any(x is None for x in [ev, epv, rv, rpv]):
            continue
        if any(x <= MIN_GMV for x in [ev, epv, rv, rpv]):
            continue
        if not _is_bare(ws_f_curr, r, rsc_n):
            continue
        ey = _val(ws_prev, r, exp_yoy) if exp_yoy else None
        ry = _val(ws_prev, r, rsc_yoy) if rsc_yoy else None
        results.append((r, label, ev, epv, rv, rpv, ey, ry))
    return results


# ─────────────────────────────────────────────────────────────────────────────
# Gap analysis — orchestrator
# ─────────────────────────────────────────────────────────────────────────────

def _run_gap_analysis(ws, new_month: str,
                      section_ends: List[str],
                      new_cols:     List[str],
                      src_path:     str) -> None:
    """
    HOW COLUMN POSITIONS ARE DETERMINED:
      new_cols[0]     = THIS month's Expert column just written by drag (e.g. T = May'26).
      section_ends[0] = PREVIOUS month's Expert column in master (e.g. S = Apr'26).
      new_cols[1]     = THIS month's RSC column just written by drag (e.g. AK = May'26).
      section_ends[1] = PREVIOUS month's RSC column in master (e.g. AJ = Apr'26).
      All four come from the pipeline contract (find_all_month_cols_in_row + drag in run()).

    HOW HEADERS ARE LABELLED:
      All column headers use `new_month` (e.g. "May'26"), from run_pipeline.py — never hardcoded.

    WHERE OUTPUT IS WRITTEN:
      ONLY to the far-right block at new_cols[-1]+1 … new_cols[-1]+7.
      Nothing adjacent to the RSC block (grows one column right every month).

    WHERE VALUES ARE READ:
      Current month (new_cols): from `ws` — the output workbook after drag (T, AK).
      Previous month (section_ends): from src_path master with data_only=True (S, AJ cached).
    """
    print(f"\n  [EM Gap] Loading cached values from {os.path.basename(src_path)} …")
    try:
        wb_v = openpyxl.load_workbook(src_path, data_only=True,  read_only=True)
        wb_f = openpyxl.load_workbook(src_path, data_only=False, read_only=True)
    except Exception as exc:
        print(f"  [EM Gap] Cannot open src file: {exc}")
        return

    ws_v = wb_v[SHEET_NAME]
    ws_f = wb_f[SHEET_NAME]

    # ── Column references — from pipeline contract (new_cols = this month) ──
    exp_old_c = CN(new_cols[0])          # THIS month's Expert (May'26 = T)
    exp_prv_c = CN(section_ends[0])      # PREVIOUS month's Expert (Apr'26 = S)
    rsc_old_c = CN(new_cols[1])          # THIS month's RSC (May'26 = AK)
    rsc_prv_c = CN(section_ends[1])      # PREVIOUS month's RSC (Apr'26 = AJ)

    if exp_prv_c <= EXPERT_START_COL:
        print("  [EM Gap] Expert block too short for MoM — skipping.")
        return

    # ── YoY base: year-ago of new_month (e.g. May'26 → May'25) ─────────────
    #   Scan HDR_ROW in src_path master for that label to find YoY columns.
    exp_month_label = new_month
    exp_yoy_c = rsc_yoy_c = None
    yoy_base_label = None
    try:
        yoy_base_label = year_ago_label_for(str(exp_month_label))
        yoy_target     = parse_month_label(yoy_base_label)
        if yoy_target:
            for c in range(EXPERT_START_COL, exp_prv_c):
                lbl = ws_v.cell(HDR_ROW, c).value
                if lbl and parse_month_label(str(lbl)) == yoy_target:
                    exp_yoy_c = c
                    break
            for c in range(exp_prv_c + 1, rsc_prv_c):
                lbl = ws_v.cell(HDR_ROW, c).value
                if lbl and parse_month_label(str(lbl)) == yoy_target:
                    rsc_yoy_c = c
                    break
    except ValueError:
        pass

    print(f"  [EM Gap] new_month   : {new_month!r}  ← label for all output headers")
    print(f"  [EM Gap]   (source   : run() parameter, from run_pipeline.py contract)")
    print(f"  [EM Gap] Expert: {CL(exp_old_c)} ({new_month}), comparing against "
          f"{CL(exp_prv_c)} (prev); YoY={CL(exp_yoy_c) if exp_yoy_c else 'N/A'} ({yoy_base_label})")
    print(f"  [EM Gap] RSC:    {CL(rsc_old_c)} ({new_month}), comparing against "
          f"{CL(rsc_prv_c)} (prev); YoY={CL(rsc_yoy_c) if rsc_yoy_c else 'N/A'}")
    print(f"  [EM Gap]   (newest from new_cols={new_cols[:2]}, prev from section_ends={section_ends[:2]})")

    # ── Category rows ───────────────────────────────────────────────────────
    rows_data = _find_category_rows(
        ws, ws, ws_v, ws_f,
        exp_old_c, exp_prv_c, rsc_old_c, rsc_prv_c, exp_yoy_c, rsc_yoy_c,
    )
    print(f"  [EM Gap] Rows found  : {len(rows_data)} with hardcoded RSC newest")
    if not rows_data:
        print("  [EM Gap] Nothing to analyse — skipping.")
        return

    # ── Far-right output columns ────────────────────────────────────────────
    # Always placed at new_cols[-1]+1 (first column after the last dragged section).
    # This shifts right by one each month alongside every EM section — never collides.
    # The previous month's first far-right column is overwritten by the section drag
    # (acceptable; the current month is always fresh and correct).
    far_n = CN(new_cols[-1]) + 1

    print(f"  [EM Gap] Far-right   : cols {CL(far_n)}–{CL(far_n+6)}")
    print(f"  [EM Gap] RSC block ends at {new_cols[1]}. "
          f"Next month grows to {CL(CN(new_cols[1])+1)}. "
          f"Nothing written at {CL(CN(new_cols[1])+1)} — no collision.")

    # ── Headers ─────────────────────────────────────────────────────────────
    col_headers = [
        f"Expert MoM% ({new_month})",
        f"RSC MoM% ({new_month})",
        f"MoM Gap pp ({new_month})",
        f"Expert YoY% ({new_month})",
        f"RSC YoY% ({new_month})",
        f"YoY Gap pp ({new_month})",
        f"RSC Suggested ({new_month})",
    ]
    for i, hdr_text in enumerate(col_headers):
        cell = ws.cell(HDR_ROW, far_n + i)
        cell.value = hdr_text
        cell.font  = BOLD
        cell.fill  = GREY_FILL

    # ── Process rows ─────────────────────────────────────────────────────────
    print()
    print(f"  {'Category':<38} {'ExpMoM%':>8} {'RscMoM%':>8} {'MoMGap':>7}"
          f" {'MoM':>5} {'ExpYoY%':>8} {'RscYoY%':>8} {'YoYGap':>7} {'YoY':>5} {'Sugg':>10}")
    print(f"  {'-'*38} {'-'*8} {'-'*8} {'-'*7} {'-'*5} {'-'*8} {'-'*8} {'-'*7} {'-'*5} {'-'*10}")

    mom_flag_count = yoy_flag_count = 0

    for row_data in rows_data:
        r, label, ev, epv, rv, rpv, ey, ry = row_data

        exp_mom  = _pct(ev,  epv)
        rsc_mom  = _pct(rv,  rpv)
        mom_gap  = (abs(exp_mom - rsc_mom)
                    if exp_mom is not None and rsc_mom is not None else None)
        mom_flag = mom_gap is not None and mom_gap > MOM_GAP_THRESHOLD

        exp_yoy_pct = _pct(ev, ey) if (ey and ey > MIN_GMV) else None
        rsc_yoy_pct = _pct(rv, ry) if (ry and ry > MIN_GMV) else None
        yoy_gap     = (abs(exp_yoy_pct - rsc_yoy_pct)
                       if exp_yoy_pct is not None and rsc_yoy_pct is not None else None)
        yoy_flag    = yoy_gap is not None and yoy_gap > YOY_GAP_THRESHOLD

        suggested = None
        if mom_flag and exp_mom is not None and rsc_mom is not None:
            mom_flag_count += 1
            suggested = suggest_adjustment_placeholder(rpv, exp_mom / 100, rsc_mom / 100)
        if yoy_flag:
            yoy_flag_count += 1

        # Write far-right values (rows aligned to actual EM row numbers)
        metrics = [exp_mom, rsc_mom, mom_gap, exp_yoy_pct, rsc_yoy_pct, yoy_gap]
        gap_cols = {2, 5}
        for i, v in enumerate(metrics):
            cell = ws.cell(r, far_n + i)
            if v is None:
                cell.value = "N/A"
            else:
                cell.value = round(v, 2)
                cell.number_format = '0.00' if i in gap_cols else '+0.00;-0.00;0.00'
                if i == 2:
                    cell.fill = RED_FILL if mom_flag else GREEN_FILL
                elif i == 5:
                    cell.fill = RED_FILL if yoy_flag else GREEN_FILL

        sg_cell = ws.cell(r, far_n + 6)
        if suggested is not None:
            sg_cell.value = round(suggested, 2)
            sg_cell.fill  = YELLOW_FILL
            comment_txt = (
                f"Category: {label}\n"
                f"Expert MoM%: {exp_mom:+.2f}%  RSC MoM%: {rsc_mom:+.2f}%  "
                f"MoM Gap: {mom_gap:.2f}pp\n"
                f"Placeholder: RSC_prev×(1+{EXPERT_WEIGHT}×Exp+{RSC_WEIGHT}×RSC)\n"
                f"Replace suggest_adjustment_placeholder() body to wire Gemini."
            )
            sg_cell.comment = Comment(comment_txt, "em_gap_analysis")
            sg_cell.comment.width  = 300
            sg_cell.comment.height = 100
        else:
            sg_cell.value = "—"

        # Terminal row
        def _f(v): return f"{v:+8.2f}" if v is not None else "     N/A"
        def _g(v): return f"{v:7.2f}"  if v is not None else "    N/A"
        mf = "YES◀" if mom_flag else "  no"
        yf = "YES◀" if yoy_flag else "  no"
        sg = f"{suggested:>10,.2f}" if suggested is not None else "         —"
        print(f"  {label[:38]:<38} {_f(exp_mom)} {_f(rsc_mom)} {_g(mom_gap)}"
              f" {mf} {_f(exp_yoy_pct)} {_f(rsc_yoy_pct)} {_g(yoy_gap)} {yf} {sg}")

    print()
    print(f"  [EM Gap] MoM flagged : {mom_flag_count}/{len(rows_data)}  (>{MOM_GAP_THRESHOLD}pp)")
    print(f"  [EM Gap] YoY flagged : {yoy_flag_count}/{len(rows_data)}  (>{YOY_GAP_THRESHOLD}pp)")
    print(f"  [EM Gap] Cols written: {CL(far_n)}–{CL(far_n+6)} only.")
    print(f"  [EM Gap] Checker rows 109/135 — columns never touched. ✓")
