"""
PHASE 0 — New_Meesho Masterfile discovery audit.

READ-ONLY. Writes nothing. Prints everything to terminal.

Usage:
    python new_meesho_masterfile_audit.py

The master workbook must NOT be open in Excel when you run this.
"""
import re
import sys
import os
from collections import defaultdict
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string as CN, get_column_letter as CL

# ── Config ────────────────────────────────────────────────────────────────────
MASTER_FILE  = "MeeshoMasterfilesampleautomation.xlsx"
SHEET_NAME   = "New_Meesho Masterfile"
MAX_ROWS     = 800   # scan ceiling
MAX_COLS     = 200   # scan ceiling for header row

MONTH_RE = re.compile(
    r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)'(\d{2})", re.I
)

# Labels that indicate a checker/validation row
CHECKER_LABELS = re.compile(r"check|checker|total check|validation", re.I)

# ── Column reference extractor ────────────────────────────────────────────────
CELL_REF_RE = re.compile(r"(\$?)([A-Z]+)(\$?)(\d+)")


def extract_col_offsets(formula: str, self_col_n: int) -> List[int]:
    """Return list of (referenced_col - self_col) for every cell ref in formula."""
    offsets = []
    for m in CELL_REF_RE.finditer(formula):
        col_n = CN(m.group(2))
        offsets.append(col_n - self_col_n)
    return offsets


def classify_row(
    row_n: int,
    cur_col_n: int,   # newest month column number
    prv_col_n: int,   # previous month column number
    ws_f,             # formula workbook sheet
    ws_v,             # data_only workbook sheet
    label: str,
) -> Tuple[str, dict]:
    """
    Returns (classification, detail_dict).
    Classifications: FORMULA_RELATIVE | FORMULA_INCONSISTENT | HARDCODED | EMPTY | CHECKER
    """
    # Checker first
    if CHECKER_LABELS.search(label):
        cur_val = ws_v.cell(row_n, cur_col_n).value
        cur_formula = ws_f.cell(row_n, cur_col_n).value
        return "CHECKER", {"label": label, "formula": cur_formula, "value": cur_val}

    cur_cell_f = ws_f.cell(row_n, cur_col_n).value
    prv_cell_f = ws_f.cell(row_n, prv_col_n).value

    # Both empty
    if cur_cell_f is None and prv_cell_f is None:
        return "EMPTY", {}

    # Current is bare number
    if isinstance(cur_cell_f, (int, float)):
        return "HARDCODED", {"value": cur_cell_f}

    # Current is formula
    if isinstance(cur_cell_f, str) and cur_cell_f.startswith("="):
        cur_offsets = sorted(set(extract_col_offsets(cur_cell_f, cur_col_n)))
        if isinstance(prv_cell_f, str) and prv_cell_f.startswith("="):
            prv_offsets = sorted(set(extract_col_offsets(prv_cell_f, prv_col_n)))
            if cur_offsets == prv_offsets:
                return "FORMULA_RELATIVE", {
                    "offsets": cur_offsets,
                    "cur_formula": cur_cell_f,
                    "prv_formula": prv_cell_f,
                }
            else:
                return "FORMULA_INCONSISTENT", {
                    "cur_offsets": cur_offsets,
                    "prv_offsets": prv_offsets,
                    "cur_formula": cur_cell_f,
                    "prv_formula": prv_cell_f,
                }
        else:
            # Previous is bare/missing but current is formula — treat as relative
            # (can't compare offsets; flag with note)
            return "FORMULA_INCONSISTENT", {
                "cur_offsets": cur_offsets,
                "prv_offsets": None,
                "note": "prev cell is not a formula",
                "cur_formula": cur_cell_f,
                "prv_formula": prv_cell_f,
            }

    # Current is None but previous has data
    if cur_cell_f is None and prv_cell_f is not None:
        return "EMPTY", {"note": "prev has data but cur is empty"}

    return "EMPTY", {}


# ── Block detection ───────────────────────────────────────────────────────────

def find_blocks(ws_v, label_col_n: int, max_rows: int, cur_col_n: int) -> List[dict]:
    """
    A 'block' starts when column A (or B) has a non-numeric, non-month
    label that acts as a section header, and the following rows have data
    in the current month column. We detect block boundaries as rows where
    the label column contains text AND the row directly above is empty or
    also a header, and the rows below have numeric data.

    Strategy: collect all rows that have a non-empty label AND have either
    a formula or number in cur_col_n. Then group consecutive runs.
    """
    # First pass: find all rows with text in col A
    blocks = []
    current_block = None

    for r in range(1, max_rows + 1):
        a_val = ws_v.cell(r, 1).value   # Column A
        b_val = ws_v.cell(r, 2).value   # Column B (sometimes used for labels)
        cur_val = ws_v.cell(r, cur_col_n).value

        label = str(a_val).strip() if a_val is not None else ""
        b_label = str(b_val).strip() if b_val is not None else ""

        # Section header: non-empty A, no value in current month col, not a number
        is_header = (
            label
            and not isinstance(a_val, (int, float))
            and cur_val is None
            and not MONTH_RE.search(label)
        )

        # Data row: something in current month col
        has_data = cur_val is not None

        if is_header and current_block:
            blocks.append(current_block)
            current_block = None

        if is_header:
            current_block = {
                "header": label,
                "header_row": r,
                "start_row": r + 1,
                "end_row": r,
                "data_rows": 0,
            }
        elif has_data and current_block:
            current_block["end_row"] = r
            current_block["data_rows"] += 1

    if current_block and current_block["data_rows"] > 0:
        blocks.append(current_block)

    return blocks


# ── Header row scanner ────────────────────────────────────────────────────────

def find_month_columns(ws_v, max_cols: int) -> Tuple[Optional[int], Optional[int], str, str]:
    """
    Scan the first 10 rows for a header row containing month labels.
    Returns (newest_col_n, prev_col_n, newest_label, prev_label).
    """
    best_row = None
    best_months = []

    for r in range(1, 15):
        months_in_row = []
        for c in range(1, max_cols + 1):
            v = ws_v.cell(r, c).value
            if v and MONTH_RE.search(str(v)):
                months_in_row.append((c, str(v).strip()))
        if len(months_in_row) > len(best_months):
            best_months = months_in_row
            best_row = r

    if len(best_months) < 2:
        return None, None, "", ""

    # Rightmost = newest, second-rightmost = previous
    newest_col_n, newest_label = best_months[-1]
    prev_col_n, prev_label     = best_months[-2]
    return newest_col_n, prev_col_n, newest_label, prev_label


# ── Row label resolver ────────────────────────────────────────────────────────

def row_label(ws_v, row_n: int) -> str:
    a = ws_v.cell(row_n, 1).value
    b = ws_v.cell(row_n, 2).value
    return str(a).strip() if a is not None else (str(b).strip() if b is not None else "")


# ── Main audit ────────────────────────────────────────────────────────────────

def main():
    if not os.path.exists(MASTER_FILE):
        sys.exit(f"ERROR: {MASTER_FILE!r} not found. Run from the project directory.")

    print(f"Loading {MASTER_FILE} (no read_only — needed for random cell access) …")
    try:
        # NOTE: read_only=True uses streaming and does NOT support ws.cell(r,c)
        # random access. Load normally — file is ~3.5 MB so this is fast enough.
        wb_v = openpyxl.load_workbook(MASTER_FILE, data_only=True)
        wb_f = openpyxl.load_workbook(MASTER_FILE, data_only=False)
    except PermissionError:
        sys.exit(
            "ERROR: File is locked. Close MeeshoMasterfilesampleautomation.xlsx "
            "in Excel first, then re-run."
        )

    if SHEET_NAME not in wb_v.sheetnames:
        sys.exit(
            f"ERROR: Sheet {SHEET_NAME!r} not found.\n"
            f"Available sheets: {wb_v.sheetnames}"
        )

    ws_v = wb_v[SHEET_NAME]
    ws_f = wb_f[SHEET_NAME]

    print(f"Sheet: {SHEET_NAME!r}")

    # ── 1. Find header row and month columns ─────────────────────────────────
    newest_col_n, prev_col_n, newest_label, prev_label = find_month_columns(ws_v, MAX_COLS)
    if not newest_col_n:
        sys.exit("ERROR: Could not find month header row in the sheet.")

    print(f"\n{'='*70}")
    print(f"HEADER ROW SCAN")
    print(f"{'='*70}")
    print(f"  Newest month  : {newest_label!r:20}  col {CL(newest_col_n)} ({newest_col_n})")
    print(f"  Previous month: {prev_label!r:20}  col {CL(prev_col_n)} ({prev_col_n})")
    print(f"  Column gap    : {newest_col_n - prev_col_n} (expected 1 if monthly)")

    # ── 2. Classify every row ─────────────────────────────────────────────────
    counts: Dict[str, int] = defaultdict(int)
    by_class: Dict[str, list] = defaultdict(list)
    offset_freq: Dict[tuple, int] = defaultdict(int)
    checker_rows = []

    for r in range(1, MAX_ROWS + 1):
        label = row_label(ws_v, r)
        cls, detail = classify_row(r, newest_col_n, prev_col_n, ws_f, ws_v, label)
        counts[cls] += 1
        by_class[cls].append((r, label, detail))
        if cls == "FORMULA_RELATIVE":
            key = tuple(detail["offsets"])
            offset_freq[key] += 1
        if cls == "CHECKER":
            checker_rows.append((r, label, detail))

    # ── 3. Block detection ───────────────────────────────────────────────────
    blocks = find_blocks(ws_v, 1, MAX_ROWS, newest_col_n)

    print(f"\n{'='*70}")
    print(f"BLOCKS DETECTED (section header → data rows)")
    print(f"{'='*70}")
    if not blocks:
        print("  (no blocks detected — sheet may use a different header convention)")
    for i, b in enumerate(blocks, 1):
        print(f"  Block {i:2d}  rows {b['start_row']:4d}–{b['end_row']:4d}"
              f"  ({b['data_rows']:3d} data rows)  header: {b['header'][:60]!r}")

    # ── 4. Row classification summary ────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"ROW CLASSIFICATION SUMMARY (rows 1–{MAX_ROWS})")
    print(f"{'='*70}")
    total = sum(counts.values())
    for cls in ["FORMULA_RELATIVE", "FORMULA_INCONSISTENT", "HARDCODED",
                "CHECKER", "EMPTY"]:
        n = counts.get(cls, 0)
        print(f"  {cls:<25}  {n:4d}  ({100*n/total:.1f}%)")
    print(f"  {'TOTAL':<25}  {total:4d}")

    # ── 5. Offset patterns for FORMULA_RELATIVE ──────────────────────────────
    print(f"\n{'='*70}")
    print(f"FORMULA_RELATIVE — OFFSET PATTERNS")
    print(f"{'='*70}")
    if not offset_freq:
        print("  (no FORMULA_RELATIVE rows found)")
    else:
        sorted_patterns = sorted(offset_freq.items(), key=lambda x: -x[1])
        print(f"  {len(sorted_patterns)} distinct pattern(s):")
        for pat, cnt in sorted_patterns[:30]:  # cap at 30 to stay readable
            print(f"    offsets={list(pat)}  count={cnt}")
        if len(sorted_patterns) > 30:
            print(f"    … and {len(sorted_patterns)-30} more distinct patterns")
        print()
        print("  NOTE: If there are many distinct patterns, each must be verified")
        print("  before Phase 1 — a 'consistent' offset is only safe if it matches")
        print("  how those rows were constructed in the prior months.")

    # ── 6. FORMULA_INCONSISTENT detail ───────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"FORMULA_INCONSISTENT rows (need manual review)")
    print(f"{'='*70}")
    if not by_class["FORMULA_INCONSISTENT"]:
        print("  (none)")
    else:
        for r, label, detail in by_class["FORMULA_INCONSISTENT"]:
            print(f"  Row {r:4d}  label={label[:40]!r}")
            print(f"         cur  offsets={detail.get('cur_offsets')}  "
                  f"formula={str(detail.get('cur_formula',''))[:80]!r}")
            print(f"         prev offsets={detail.get('prv_offsets')}  "
                  f"formula={str(detail.get('prv_formula',''))[:80]!r}")
            if detail.get('note'):
                print(f"         note: {detail['note']}")

    # ── 7. HARDCODED rows ────────────────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"HARDCODED rows (bare numbers — analyst assumptions)")
    print(f"{'='*70}")
    if not by_class["HARDCODED"]:
        print("  (none)")
    else:
        for r, label, detail in by_class["HARDCODED"]:
            print(f"  Row {r:4d}  label={label[:50]!r}  value={detail.get('value')}")

    # ── 8. Checker row inventory ─────────────────────────────────────────────
    print(f"\n{'='*70}")
    print(f"CHECKER ROWS — full inventory (data_only values)")
    print(f"{'='*70}")
    if not checker_rows:
        print("  (none found)")
        print("  NOTE: Checker rows may use a different label convention in this sheet.")
        print("  Scanning column A for rows containing numeric checker formulas …")
        # Fallback: scan for rows whose current-month cell is a formula referencing
        # many columns (typical checker: =SUM, =A-B, etc.) and prev cell is also formula
        for r in range(1, MAX_ROWS + 1):
            cur_f = ws_f.cell(r, newest_col_n).value
            if isinstance(cur_f, str) and cur_f.startswith("=") and (
                "SUM" in cur_f.upper() or "-" in cur_f
            ):
                cur_v = ws_v.cell(r, newest_col_n).value
                lbl = row_label(ws_v, r)
                if cur_v is not None and float(cur_v) == 0:
                    print(f"  Row {r:4d}  label={lbl[:40]!r}  formula={cur_f[:80]!r}  value={cur_v}")
    else:
        for r, label, detail in checker_rows:
            print(f"  Row {r:4d}  label={label[:50]!r}")
            print(f"         formula={str(detail.get('formula',''))[:80]!r}")
            print(f"         value  ={detail.get('value')!r}")

    # ── 9. Sample formulas from first few data rows ──────────────────────────
    print(f"\n{'='*70}")
    print(f"SAMPLE — first 8 FORMULA_RELATIVE rows (spot-check offset logic)")
    print(f"{'='*70}")
    shown = 0
    for r, label, detail in by_class["FORMULA_RELATIVE"]:
        if shown >= 8:
            break
        cur_f = detail.get("cur_formula", "")
        prv_f = detail.get("prv_formula", "")
        offsets = detail.get("offsets", [])
        print(f"  Row {r:4d}  label={label[:35]!r}  offsets={offsets}")
        print(f"         cur ({CL(newest_col_n)}): {str(cur_f)[:90]!r}")
        print(f"         prv ({CL(prev_col_n)}): {str(prv_f)[:90]!r}")
        shown += 1

    print(f"\n{'='*70}")
    print(f"PHASE 0 COMPLETE — review before proceeding to Phase 1")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
