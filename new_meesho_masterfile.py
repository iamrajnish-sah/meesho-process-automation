"""
STEP 7 — New_Meesho Masterfile
Drags the newest data column and the newest MoM% column forward by one month.
Also extends the Contribution mini-timeline (insert May before YoY at BW).
"""
import random
import re
from typing import Dict, List, Optional, Tuple

from openpyxl.utils import column_index_from_string as CN, get_column_letter as CL

from shared.cancel_generation import (
    generate_cancel_pct,
    hist_mom_swings_pp_from_col_n,
)
from shared.contracts import EMContract, FKContract, MPVContract, NMMContract
from shared.formula_utils import remap_formula_refs, shift_formula, col_num, col_str
from shared.month_utils import parse_month_label
from openpyxl import load_workbook

from shared.nmm_formula_eval import contrib_formula_rows_share_sum
from shared.style_utils import YELLOW_FILL, add_comment, copy_cell_style
from shared.trend_hardcode import (
    generate_pp_bounded,
    is_bare,
    write_hardcoded_cell,
)

SHEET_NAME  = "New_Meesho Masterfile"
HEADER_ROW  = 8
DATA_ROW_START = 9
DATA_ROW_END   = 800       # generous upper bound; empty rows are skipped

# Phase 0 audit: MoM% at EP row 9 = =BB9/BA9-1  →  EP(146) − BB(54) = 92.
# The old offset 91 landed one column right (BC), which is always empty on the
# master — a silent skip, not an exception.
DATA_MOM_OFFSET = 92

# Main metrics timeline (J…BB) ends before the Contribution mini-timeline (BG+).
DATA_METRICS_COL_START = CN("J")    # 10
DATA_METRICS_COL_END = CN("BE")     # 57 — cols 55–58 (BC–BF) are structural gap
MOM_SECTION_COL_START = CN("EG")    # 137 — MoM% headers begin here

# Regex to detect cross-references to the Revision sheet (not in pipeline scope)
_REVISION_RE = re.compile(r"Revision!", re.IGNORECASE)

# Contribution section hardcoded / generated rows
NMM_CONTRIB_HARDCODE_ROWS = (65, 66, 67)
NMM_CONTRIB_FORMULA_ROWS = (68, 69, 70, 71, 72)  # =SUM with row 73
NMM_CONTRIB_TOTAL_ROW = 73
CONTRIB_TOTAL_TARGET = 1.0  # 100% stored as unity fraction (matches BV73 ≈ 1.0)
NMM_CAT_CANCEL_ROWS = tuple(range(52, 60))
NMM_SKIP_ROWS = frozenset({352})
NMM_CONTRIB_PP_MONTHS = 6
NMM_L2_ROW_START = 292
NMM_L2_ROW_END = 360


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def _is_clean_month_header(value) -> Optional[str]:
    if not isinstance(value, str):
        return None
    label = value.strip()
    try:
        parse_month_label(label)
    except Exception:
        return None
    if " " in label or label.count("'") != 1:
        return None
    return label


def _find_data_metrics_end_col(ws) -> int:
    """Rightmost clean month header in the main metrics block (J…BE).

    Does NOT scan Contribution (BG+) or MoM% (EG+) — those are separate
    sub-blocks with their own timelines (see rows 65–67 audit).
    """
    for c in range(DATA_METRICS_COL_END, DATA_METRICS_COL_START - 1, -1):
        label = _is_clean_month_header(ws.cell(HEADER_ROW, c).value)
        if label is not None:
            return c
    raise RuntimeError(
        f"[NMM] No month header found in metrics block "
        f"{CL(DATA_METRICS_COL_START)}…{CL(DATA_METRICS_COL_END)} row {HEADER_ROW}."
    )


def _find_mom_section_end_col(ws, max_cols: int = 200) -> int:
    """Rightmost clean month header in the MoM% block (EG…)."""
    for c in range(max_cols, MOM_SECTION_COL_START - 1, -1):
        label = _is_clean_month_header(ws.cell(HEADER_ROW, c).value)
        if label is not None:
            return c
    raise RuntimeError(
        f"[NMM] No month header found in MoM% block from "
        f"{CL(MOM_SECTION_COL_START)} row {HEADER_ROW}."
    )


def _count_filled_rows(ws, col_n: int) -> int:
    return sum(
        1 for r in range(DATA_ROW_START, DATA_ROW_END + 1)
        if ws.cell(r, col_n).value is not None
    )


def _find_section_end_cols(ws, max_cols: int = 200) -> Tuple[int, int]:
    """Return (data_col_n, mom_col_n) for the metrics + MoM% pair.

    Uses scoped scans (metrics J…BE, MoM EG…) instead of a global row-8 scan,
    which incorrectly picked MoM labels and derived empty BC via offset 91.
    """
    data_col_n = _find_data_metrics_end_col(ws)
    mom_col_n = _find_mom_section_end_col(ws, max_cols=max_cols)
    offset = mom_col_n - data_col_n

    data_label = _is_clean_month_header(ws.cell(HEADER_ROW, data_col_n).value)
    mom_label = _is_clean_month_header(ws.cell(HEADER_ROW, mom_col_n).value)

    if offset != DATA_MOM_OFFSET:
        raise RuntimeError(
            f"[NMM] Metrics/MoM offset mismatch: {CL(mom_col_n)}({mom_label!r}) − "
            f"{CL(data_col_n)}({data_label!r}) = {offset}, expected {DATA_MOM_OFFSET}. "
            "Sheet layout may have changed — re-run new_meesho_masterfile_audit.py."
        )

    filled = _count_filled_rows(ws, data_col_n)
    if filled == 0:
        raise RuntimeError(
            f"[NMM] Metrics anchor {CL(data_col_n)} ({data_label!r}) has 0 data rows — "
            "refusing to drag from an empty column."
        )

    print(
        f"  [NMM] Anchor validated: metrics {CL(data_col_n)} ({data_label!r}, "
        f"{filled} rows) <-> MoM {CL(mom_col_n)} ({mom_label!r}), offset={offset}"
    )
    return data_col_n, mom_col_n


def _find_contribution_cols(ws) -> Tuple[int, int]:
    """Return (prev_month_col_n, yoy_col_n) for the Contribution mini-timeline."""
    for c in range(CN("BG"), 220):
        v = ws.cell(HEADER_ROW, c).value
        if isinstance(v, str) and v.strip() == "YoY":
            return c - 1, c
    raise RuntimeError(
        "[NMM] Contribution 'YoY' column not found in row 8 (expected before BW)."
    )


def _drag_contribution_section(
    ws, new_month: str,
) -> Tuple[str, str, int]:
    """Insert new month before YoY, drag prev month → new column."""
    prev_n, yoy_n = _find_contribution_cols(ws)
    prev_col = CL(prev_n)
    prev_label = ws.cell(HEADER_ROW, prev_n).value
    print(f"  [NMM Contrib] Confirmed Apr column: {prev_col} ({prev_label!r})")
    print(f"  [NMM Contrib] Inserting {new_month} before YoY at {CL(yoy_n)} …")

    ws.insert_cols(yoy_n)
    new_n = yoy_n
    new_col = CL(new_n)
    new_hdr = ws.cell(HEADER_ROW, new_n)
    new_hdr.value = new_month
    copy_cell_style(ws.cell(HEADER_ROW, prev_n), new_hdr)

    count = 0
    delta = new_n - prev_n
    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        src = ws.cell(r, prev_n)
        if src.value is None:
            continue
        dst = ws.cell(r, new_n)
        if isinstance(src.value, str) and src.value.startswith("="):
            dst.value = shift_formula(src.value, delta)
            copy_cell_style(src, dst)
        elif isinstance(src.value, (int, float)):
            dst.value = src.value
            copy_cell_style(src, dst)
        else:
            dst.value = src.value
            copy_cell_style(src, dst)
        count += 1

    print(
        f"  [NMM Contrib] Dragged {count} cells {prev_col}→{new_col} "
        f"(YoY now at {CL(yoy_n + 1)})"
    )
    return prev_col, new_col, count


def _nmm_row_label(ws, row: int) -> str:
    return str(ws.cell(row, 3).value or ws.cell(row, 2).value or f"row{row}")[:38]


def _load_apr_contrib_formula_shares(master_path: Optional[str]) -> Optional[float]:
    """Sum of BV68:BV72 from master (data_only) — Apr formula-row shares."""
    if not master_path:
        return None
    try:
        wb_ro = load_workbook(master_path, data_only=True, read_only=True)
        ws = wb_ro[SHEET_NAME]
        total = 0.0
        ok = False
        for row in NMM_CONTRIB_FORMULA_ROWS:
            v = ws.cell(row, CN("BV")).value
            if isinstance(v, (int, float)):
                total += float(v)
                ok = True
        wb_ro.close()
        return total if ok else None
    except Exception:
        return None


def _formula_rows_share_sum(
    wb,
    metrics_col: str,
    master_path: Optional[str],
) -> Tuple[float, str]:
    """Return (share_sum, source_label) for rows 68-72 feeding BW73."""
    share = contrib_formula_rows_share_sum(
        wb, SHEET_NAME, metrics_col, NMM_CONTRIB_FORMULA_ROWS,
    )
    if share is not None:
        return share, f"evaluated {metrics_col}/row64"

    apr_share = _load_apr_contrib_formula_shares(master_path)
    if apr_share is not None:
        return apr_share, "Apr cached shares from master (BV68:BV72)"

    total = 0.0
    for row in NMM_CONTRIB_FORMULA_ROWS:
        v = wb.cell(row, CN("BV")).value
        if isinstance(v, (int, float)):
            total += float(v)
    return total, "Apr numeric fallback from BV68:BV72"


def _fill_contrib_rows_657_balanced(
    wb,
    ws,
    new_month: str,
    prev_data_col: str,
    new_data_col: str,
    prev_contrib_col: str,
    new_contrib_col: str,
    rng: random.Random,
    master_path: Optional[str] = None,
) -> None:
    """Generate rows 65-67 with ±1pp trend, floor at 0, scale to close BW73 to 100%."""
    prev_contrib_n = CN(prev_contrib_col)
    new_contrib_n = CN(new_contrib_col)

    generated: Dict[int, Tuple[float, dict]] = {}
    for row in NMM_CONTRIB_HARDCODE_ROWS:
        if not is_bare(ws, row, prev_contrib_n):
            continue
        new_val, meta = generate_pp_bounded(
            ws, row, prev_contrib_n,
            months=NMM_CONTRIB_PP_MONTHS, rng=rng, floor=0.0,
        )
        if new_val is None:
            continue
        generated[row] = (max(0.0, new_val), meta)

    if not generated:
        print("  [NMM Contrib 65-67] No bare rows to generate — skipping balance.")
        return

    raw_sum = sum(v for v, _ in generated.values())
    print(f"  [NMM Contrib 65-67] Generated raw sum (rows 65-67 only): {raw_sum:.6g}")

    formula_share, share_src = _formula_rows_share_sum(
        wb, new_data_col, master_path,
    )
    print(
        f"  [NMM Contrib 65-67] Formula rows 68-72 share ({share_src}): "
        f"{formula_share:.6g}"
    )
    print(
        f"  [NMM Contrib] {new_contrib_col}{NMM_CONTRIB_TOTAL_ROW} components: "
        f"rows 65-67 + rows 68-72 (formula) = SUM({new_contrib_col}65:{new_contrib_col}72)"
    )

    before_total = raw_sum + formula_share
    print(
        f"  [NMM Contrib] {new_contrib_col}{NMM_CONTRIB_TOTAL_ROW} before adjust "
        f"(65-67 raw + 68-72): {before_total:.6g} "
        f"({before_total * 100:.4f}%)"
    )

    target_657 = CONTRIB_TOTAL_TARGET - formula_share
    target_657 = max(0.0, target_657)
    current_657 = raw_sum
    if current_657 <= 0:
        each = target_657 / len(generated)
        scaled = {row: each for row in generated}
    else:
        scale = target_657 / current_657
        scaled = {row: max(0.0, v * scale) for row, (v, _) in generated.items()}

    keys = sorted(scaled.keys())
    residual = CONTRIB_TOTAL_TARGET - formula_share - sum(scaled.values())
    if keys and abs(residual) > 1e-12:
        scaled[keys[-1]] = max(0.0, scaled[keys[-1]] + residual)

    after_total = sum(scaled.values()) + formula_share
    print(
        f"  [NMM Contrib] {new_contrib_col}{NMM_CONTRIB_TOTAL_ROW} after adjust: "
        f"{after_total:.6g} ({after_total * 100:.6f}%) "
        f"[rows 73 = SUM({new_contrib_col}65:{new_contrib_col}72)]"
    )

    print(
        f"  [NMM Contrib 65-67] {'Row':>4} {'Label':38} {'Raw':>12} {'Final':>12}"
    )
    for row in NMM_CONTRIB_HARDCODE_ROWS:
        if row not in generated:
            continue
        raw_v, meta = generated[row]
        final_v = scaled[row]
        write_hardcoded_cell(
            ws, row, new_contrib_n, final_v,
            tag="NMM Contrib", prev_col=prev_contrib_col,
            new_col=new_contrib_col, meta=meta,
        )
        print(
            f"  {row:>4} {_nmm_row_label(ws, row):38} "
            f"{raw_v:12.6g} {final_v:12.6g}"
        )


def _fill_nmm_hardcoded(
    wb,
    ws,
    new_month: str,
    prev_data_col: str,
    new_data_col: str,
    prev_contrib_col: str,
    new_contrib_col: str,
    master_path: Optional[str] = None,
) -> None:
    """Generate May hardcoded cells: contrib 65-67, cat cancel 52-59, L2 contrib bare."""
    prev_data_n = CN(prev_data_col)
    new_data_n = CN(new_data_col)
    prev_contrib_n = CN(prev_contrib_col)
    new_contrib_n = CN(new_contrib_col)

    print(
        f"\n  [NMM Gen] Contribution May column: {new_contrib_col} "
        f"(from {prev_contrib_col})"
    )
    print(
        f"  [NMM Gen] {'Row':>4} {'Label':38} {'Prev':>12} {'May':>12} "
        f"{'Δpp':>8} {'Trend':>5} {'OK':>4}"
    )

    rng_c = random.Random(f"nmm-contrib|{new_month}")
    rng_x = random.Random(f"nmm-cancel|{new_month}")

    def _print_row(row: int, prev_v: float, new_v: float, meta: dict) -> None:
        ok = "yes" if meta.get("followed_trend", True) else "no"
        print(
            f"  {row:>4} {_nmm_row_label(ws, row):38} "
            f"{prev_v:12.6g} {new_v:12.6g} "
            f"{meta.get('pp_change', meta.get('delta_pp', 0)):+8.3f} "
            f"{meta.get('trend_dir', meta.get('dir', '')):>5} {ok:>4}"
        )

    _fill_contrib_rows_657_balanced(
        wb, ws, new_month,
        prev_data_col, new_data_col,
        prev_contrib_col, new_contrib_col,
        rng_c,
        master_path=master_path,
    )

    for row in NMM_CAT_CANCEL_ROWS:
        if not is_bare(ws, row, prev_data_n) or not is_bare(ws, row, new_data_n):
            continue
        try:
            apr_pct = float(ws.cell(row, prev_data_n).value)
        except (TypeError, ValueError):
            continue
        deltas = hist_mom_swings_pp_from_col_n(ws, row, prev_data_n)
        may_pct, delta_pp, dir_s, cap_pp = generate_cancel_pct(
            apr_pct, deltas, rng_x,
        )
        ws.cell(row, new_data_n).value = may_pct
        ws.cell(row, new_data_n).fill = YELLOW_FILL
        add_comment(
            ws, f"{new_data_col}{row}",
            f"[NMM Cancel] Generated {new_month} from {prev_data_col}{row}.\n"
            f"Apr={apr_pct:.4%} → May={may_pct:.4%} (Δ={delta_pp:+.3f}pp, "
            f"cap={cap_pp:.3f}pp, trend={dir_s}).",
        )
        _print_row(row, apr_pct, may_pct, {
            "pp_change": delta_pp,
            "trend_dir": dir_s,
            "followed_trend": True,
        })

    for row in range(NMM_L2_ROW_START, NMM_L2_ROW_END + 1):
        if row in NMM_SKIP_ROWS or row in NMM_CONTRIB_HARDCODE_ROWS:
            continue
        if not is_bare(ws, row, prev_contrib_n) or not is_bare(ws, row, new_contrib_n):
            continue
        new_val, meta = generate_pp_bounded(
            ws, row, prev_contrib_n,
            months=NMM_CONTRIB_PP_MONTHS, rng=rng_c,
        )
        if new_val is None:
            continue
        write_hardcoded_cell(
            ws, row, new_contrib_n, new_val,
            tag="NMM L2", prev_col=prev_contrib_col,
            new_col=new_contrib_col, meta=meta,
        )
        _print_row(row, meta["prev"], new_val, meta)


def _build_col_map(
    prev_data_col: str,
    new_data_col: str,
    mpv_contract: MPVContract,
    em_contract: EMContract,
    fk_contract: FKContract,
    extra: Optional[Dict[str, str]] = None,
) -> Dict[str, str]:
    """Build the explicit column-letter remap for the data section drag.

    Each entry maps the Apr'26 (prev) column letter to the May'26 (new)
    column letter for every cross-sheet reference that appears in the
    data column formulas.
    """
    col_map: Dict[str, str] = {}

    # NMM data column itself: BB → BC
    col_map[prev_data_col] = new_data_col

    # Meesho Prelim View: Expert prev → new, RSC prev → new
    col_map[mpv_contract["expert_prev_col"]] = mpv_contract["expert_col"]
    col_map[mpv_contract["rsc_prev_col"]]    = mpv_contract["rsc_col"]

    # Error Margin: each new col's prev = new_col − 1 column
    for nc in em_contract.get("section_new_cols", []):
        prev = col_str(col_num(nc) - 1)
        col_map[prev] = nc

    # FK Prelim View: same derivation
    for nc in fk_contract.get("section_new_cols", []):
        prev = col_str(col_num(nc) - 1)
        col_map[prev] = nc

    if extra:
        col_map.update(extra)

    return col_map


def _drag_data_col(ws, prev_col: str, new_col: str, col_map: Dict[str, str]) -> int:
    """Copy formulas/values from prev_col to new_col applying col_map remap.

    Returns count of cells written.
    """
    prev_n = col_num(prev_col)
    new_n  = col_num(new_col)
    count  = 0

    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        src_cell = ws.cell(r, prev_n)
        v = src_cell.value
        if v is None:
            continue

        dst_cell = ws.cell(r, new_n)

        if isinstance(v, str) and v.startswith("="):
            if _REVISION_RE.search(v):
                # Revision sheet is outside the pipeline — flag for manual entry
                add_comment(
                    ws, f"{new_col}{r}",
                    f"[NMM] Formula references 'Revision' sheet — NOT auto-generated.\n"
                    f"Original ({prev_col}{r}): {v}\n"
                    "Update this cell manually after running the pipeline.",
                )
                count += 1
                continue
            # Remap all known month-column references using placeholder protection.
            new_formula = remap_formula_refs(v, col_map, col_shift=0)
            dst_cell.value = new_formula
            copy_cell_style(src_cell, dst_cell)

        elif isinstance(v, (int, float)):
            # Hardcoded analyst assumption — copy forward, flag for review
            dst_cell.value = v
            copy_cell_style(src_cell, dst_cell)
            dst_cell.fill = YELLOW_FILL
            add_comment(
                ws, f"{new_col}{r}",
                f"[NMM] Hardcoded assumption carried forward from {prev_col}{r}.\n"
                "Review and update this value for the new month.",
            )

        else:
            # Plain string (non-formula), e.g. a text label
            dst_cell.value = v
            copy_cell_style(src_cell, dst_cell)

        count += 1

    return count


def _drag_mom_col(ws, prev_col: str, new_col: str) -> int:
    """Drag MoM% column forward using shift_formula(+1).

    MoM% formulas have the pattern =BB{r}/BA{r}-1 and only ever reference
    two adjacent data-section columns — shift by +1 is always correct.
    Returns count of cells written.
    """
    prev_n = col_num(prev_col)
    new_n  = col_num(new_col)
    delta  = new_n - prev_n
    count  = 0

    for r in range(DATA_ROW_START, DATA_ROW_END + 1):
        src_cell = ws.cell(r, prev_n)
        v = src_cell.value
        if v is None:
            continue
        if isinstance(v, str) and v.startswith("="):
            dst_cell = ws.cell(r, new_n)
            dst_cell.value = shift_formula(v, delta)
            copy_cell_style(src_cell, dst_cell)
            count += 1

    return count


# ──────────────────────────────────────────────────────────────────────────────
# Public entry point
# ──────────────────────────────────────────────────────────────────────────────

def run(
    wb,
    new_month: str,
    mpv_contract: MPVContract,
    em_contract: EMContract,
    fk_contract: FKContract,
    dry_run: bool = False,
    src_path: Optional[str] = None,
) -> NMMContract:
    """Step 7 — drag New_Meesho Masterfile forward one month.

    Args:
        wb           : open openpyxl workbook (already updated by steps 1-6)
        new_month    : month label string, e.g. "May'26"
        mpv_contract : contract returned by meesho_prelim.run()
        em_contract  : contract returned by error_margin.run()
        fk_contract  : contract returned by fk_prelim.run()
        dry_run      : if True, count only, do not write

    Returns NMMContract with new_data_col and new_mom_col.
    """
    ws = wb[SHEET_NAME]
    print(f"\n  [NMM] Processing New_Meesho Masterfile …")

    # 1. Locate the two newest month columns in the header row
    data_col_n, mom_col_n = _find_section_end_cols(ws)
    prev_data_col = col_str(data_col_n)      # e.g. "BB"
    new_data_col  = col_str(data_col_n + 1)  # e.g. "BC"
    prev_mom_col  = col_str(mom_col_n)       # e.g. "EO"
    new_mom_col   = col_str(mom_col_n + 1)   # e.g. "EP"

    prev_data_label = ws.cell(HEADER_ROW, data_col_n).value
    prev_mom_label  = ws.cell(HEADER_ROW, mom_col_n).value
    print(f"  [NMM] Data section  : {prev_data_col} ({prev_data_label}) → {new_data_col}")
    print(f"  [NMM] MoM%  section : {prev_mom_col}  ({prev_mom_label})  → {new_mom_col}")

    if dry_run:
        return NMMContract(new_data_col=new_data_col, new_mom_col=new_mom_col)

    # 2. Contribution section: insert May before YoY, drag BV → BW
    prev_contrib_col, new_contrib_col, _ = _drag_contribution_section(ws, new_month)

    # 3. Write headers for the new data / MoM columns
    new_data_hdr = ws.cell(HEADER_ROW, data_col_n + 1)
    new_mom_hdr  = ws.cell(HEADER_ROW, mom_col_n + 1)
    new_data_hdr.value = new_month
    new_mom_hdr.value  = new_month
    copy_cell_style(ws.cell(HEADER_ROW, data_col_n), new_data_hdr)
    copy_cell_style(ws.cell(HEADER_ROW, mom_col_n),  new_mom_hdr)

    # 4. Build column remap (includes Contribution BV → BW)
    col_map = _build_col_map(
        prev_data_col, new_data_col,
        mpv_contract, em_contract, fk_contract,
        extra={prev_contrib_col: new_contrib_col},
    )
    print(f"  [NMM] col_map ({len(col_map)} entries): "
          + ", ".join(f"{k}→{v}" for k, v in sorted(col_map.items())))

    # 5. Drag data column
    data_count = _drag_data_col(ws, prev_data_col, new_data_col, col_map)
    if data_count == 0:
        raise RuntimeError(
            f"[NMM] Data drag wrote 0 cells ({prev_data_col}→{new_data_col}). "
            f"Metrics anchor {prev_data_col} ({prev_data_label!r}) is empty — "
            "check _find_data_metrics_end_col."
        )

    # 6. Drag MoM% column
    mom_count = _drag_mom_col(ws, prev_mom_col, new_mom_col)

    # 7. Generate hardcoded May values (contrib, category cancel, L2)
    _fill_nmm_hardcoded(
        wb, ws, new_month,
        prev_data_col, new_data_col,
        prev_contrib_col, new_contrib_col,
        master_path=src_path,
    )

    print(
        f"  [NMM] Wrote {data_count} data cells ({prev_data_col}→{new_data_col}), "
        f"{mom_count} MoM% cells ({prev_mom_col}→{new_mom_col})"
    )

    return NMMContract(new_data_col=new_data_col, new_mom_col=new_mom_col)
