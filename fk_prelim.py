"""
STEP 5 — FK Prelim View
Drags all parallel section end-columns forward.
Reads MPVContract from upstream for #REF! annotation context.
Returns FKContract listing new columns written.
"""
import random
from typing import Optional
from openpyxl.comments import Comment as XLComment
from openpyxl.utils import column_index_from_string as CN

from shared.anchors import SHEET_ANCHORS, find_all_month_cols_in_row
from shared.month_utils import anchor_search_token
from shared.contracts import FKContract, MPVContract
from shared.fk_manual_inputs import (
    FKManualInputs,
    FK_MANUAL_ROWS,
    resolve_fk_manual_inputs,
)
from shared.formula_utils import drag_column_forward, next_col
from shared.style_utils import copy_cell_style, YELLOW_FILL, add_comment
from shared.trend_hardcode import generate_pp_bounded, is_bare, write_hardcoded_cell

FK_MONTH_HDR_ROW  = 6
FK_DATA_ROW_START = 7
FK_DATA_ROW_END   = 35

FK_ASSUMPTION_HIST_MONTHS = 6
FK_ROW31_FULL_BARE = 31
FK_ROW27_NEWEST  = 27


def _fix_fk_ref_errors(ws, mpv_expert_col: str) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "#REF!" in cell.value:
                try:
                    old = cell.comment.text if cell.comment else ""
                    cell.comment = XLComment(
                        f"⚠ #REF! needs manual fix. Original: {cell.value}. {old}",
                        "MeeshoBot",
                    )
                except Exception:
                    pass


def run(
    wb,
    new_month: str,
    mpv_contract: MPVContract,
    dry_run: bool,
    *,
    fk_manual: Optional[FKManualInputs] = None,
    interactive: bool = True,
) -> FKContract:
    """Entry point for Step 5."""
    ws = wb["FK Prelim View"]
    anchor = SHEET_ANCHORS["FK Prelim View"]
    print(f"\n  [FK] Processing FK Prelim View …")

    search = anchor_search_token(new_month)
    section_ends = find_all_month_cols_in_row(ws, anchor["header_row"], search)
    if not section_ends:
        print(f"  [FK] WARNING: No sections ending with {search!r}. Skipping.")
        return FKContract(section_new_cols=[])

    print(f"  [FK] Found {len(section_ends)} section(s): {section_ends}")
    new_cols: list = []
    for last_col in section_ends:
        new_col = next_col(last_col)
        new_cols.append(new_col)
        if dry_run:
            print(f"  [DRY-RUN] FK section: {last_col}→{new_col}")
            continue
        ws[f"{new_col}{FK_MONTH_HDR_ROW}"].value = new_month
        copy_cell_style(ws[f"{last_col}{FK_MONTH_HDR_ROW}"],
                        ws[f"{new_col}{FK_MONTH_HDR_ROW}"])
        n = drag_column_forward(ws, last_col, new_col,
                                FK_DATA_ROW_START, FK_DATA_ROW_END, dry_run=dry_run)
        print(f"  [FK] ✓ Section {last_col}→{new_col}: {n} cells.")

    if not dry_run:
        _fix_fk_ref_errors(ws, mpv_contract["expert_col"])
        try:
            _fill_fk_manual_inputs(
                ws, new_month, new_cols[0], fk_manual, interactive=interactive,
            )
        except Exception as exc:
            print(f"  [FK Manual] ⚠ Manual ASP/AoV/Cancel fill skipped: {exc}")
        try:
            _fill_fk_hardcoded(ws, new_month, section_ends, new_cols)
        except Exception as exc:
            print(f"  [FK Assumption] ⚠ FK hardcoded fill skipped: {exc}")

    return FKContract(section_new_cols=new_cols)


def _fill_fk_manual_inputs(
    ws,
    new_month: str,
    expert_new_col: str,
    fk_manual: Optional[FKManualInputs],
    *,
    interactive: bool,
) -> None:
    """Write user-supplied ASP / AoV / Cancellation % into the newest Expert column."""
    inputs = resolve_fk_manual_inputs(
        new_month, fk_manual, interactive=interactive,
    )
    values = {
        FK_MANUAL_ROWS["asp"]: inputs.asp,
        FK_MANUAL_ROWS["aov"]: inputs.aov,
        FK_MANUAL_ROWS["cancellation_pct"]: inputs.cancellation_pct,
    }
    print(f"\n  [FK Manual] Writing user values to {expert_new_col} "
          f"(rows 7/8/11) for {new_month}:")
    for label, row in (
        ("ASP", FK_MANUAL_ROWS["asp"]),
        ("AoV", FK_MANUAL_ROWS["aov"]),
        ("Cancellation %", FK_MANUAL_ROWS["cancellation_pct"]),
    ):
        val = values[row]
        cell = ws[f"{expert_new_col}{row}"]
        cell.value = val
        cell.fill = YELLOW_FILL
        add_comment(
            ws, f"{expert_new_col}{row}",
            f"[FK Manual] User-supplied {label} for {new_month}.\n"
            f"Not auto-generated — entered via pipeline prompt / web config.",
        )
        disp = f"{val:.4%}" if row == FK_MANUAL_ROWS["cancellation_pct"] else f"{val:.4f}"
        print(f"    {label}: {disp}")


def _fk_row_label(ws, row: int) -> str:
    return str(ws.cell(row, 2).value or ws.cell(row, 1).value or f"row{row}")[:38]


def _fill_fk_hardcoded(
    ws,
    new_month: str,
    section_ends: list,
    new_cols: list,
) -> None:
    """Generate May hardcoded cells for FK rows 27 (newest-only) and 31 (full bare)."""
    # Row 27/31 live in the section ending at BC (4th section in Apr'26 master).
    target_prev = "BC"
    if target_prev not in section_ends:
        print("  [FK Assumption] BC section not found — skipping.")
        return

    idx = section_ends.index(target_prev)
    prev_col = section_ends[idx]
    new_col  = new_cols[idx]
    prev_n   = CN(prev_col)
    new_n    = CN(new_col)

    print(f"\n  [FK Assumption] Section {prev_col} → {new_col} "
          f"(±1pp MoM cap, {FK_ASSUMPTION_HIST_MONTHS}m trend):")
    print(f"  {'Row':>4} {'Label':38} {'Prev':>12} {'May':>12} "
          f"{'Δpp':>8} {'Trend':>5} {'OK':>4}")

    rng = random.Random(f"fk-assumption|{new_month}")
    written = 0

    for row, mode in ((FK_ROW31_FULL_BARE, "full"), (FK_ROW27_NEWEST, "newest")):
        if not is_bare(ws, row, prev_n):
            continue
        if mode == "newest" and is_bare(ws, row, prev_n - 1):
            continue
        if not is_bare(ws, row, new_n):
            continue

        new_val, meta = generate_pp_bounded(
            ws, row, prev_n,
            months=FK_ASSUMPTION_HIST_MONTHS,
            rng=rng,
        )
        if new_val is None or not meta:
            print(f"  {row:>4} {_fk_row_label(ws, row):38}  SKIP — no prior value")
            continue

        write_hardcoded_cell(
            ws, row, new_n, new_val,
            tag="FK Assumption",
            prev_col=prev_col,
            new_col=new_col,
            meta=meta,
        )
        ok = "yes" if meta["followed_trend"] else "no"
        print(
            f"  {row:>4} {_fk_row_label(ws, row):38} "
            f"{meta['prev']:12.6g} {new_val:12.6g} "
            f"{meta['pp_change']:+8.3f} {meta['trend_dir']:>5} {ok:>4}"
        )
        written += 1

    print(f"  [FK Assumption] Hardcoded rows written: {written}")
