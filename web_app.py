#!/usr/bin/env python3
"""
Local Flask UI for the Meesho monthly update pipeline.

Run (after installing requirements-web.txt):
    python web_app.py
    → http://127.0.0.1:8080
"""
import os
import sys
import traceback

print("Loading Meesho web UI …", flush=True)

from typing import Optional

from flask import Flask, flash, redirect, render_template, request, send_file, session, url_for

import gemini_ingest
import web_uploads
from shared.month_utils import detect_suggested_new_month
from shared.workbook_io import load_workbook_safe

# Heavy pipeline imports deferred until first pipeline run (faster, clearer startup).

if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

app = Flask(__name__)
# Stable secret for local dev (sessions only; not security-critical on localhost).
app.secret_key = os.environ.get("MEESHO_FLASK_SECRET", "meesho-local-dev-secret-change-me")

DEFAULT_PORT = int(os.environ.get("MEESHO_WEB_PORT", "8080"))
DEFAULT_HOST = os.environ.get("MEESHO_WEB_HOST", "127.0.0.1")

STAGES = [
    {
        "id": "raw_data",
        "label": "Through Raw Data",
        "desc": "Raw Working Sheet and Raw Data sheets only.",
        "css": "stage-blue",
        "download_name": "Meesho through_raw_data output.xlsx",
    },
    {
        "id": "error_margin",
        "label": "Through Error Margin",
        "desc": "Stops after Meesho Prelim View and Error Margin.",
        "css": "stage-blue",
        "download_name": "Meesho through_error_margin output.xlsx",
    },
    {
        "id": "client_prelim",
        "label": "Through Client Prelim",
        "desc": "Full six-step pipeline including Client Prelim.",
        "css": "stage-blue",
        "download_name": "Meesho through_client_prelim output.xlsx",
    },
    {
        "id": "full_master",
        "label": "Full Master",
        "desc": "Final deliverable. New_Meesho Masterfile is included unchanged.",
        "css": "stage-green",
        "download_name": "Meesho full_master output.xlsx",
    },
]


def _ensure_session() -> str:
    if "upload_id" not in session:
        session["upload_id"] = web_uploads.new_session_id()
    return session["upload_id"]


def _parse_total_gmv(raw: str) -> Optional[float]:
    raw = (raw or "").strip().replace(",", "")
    if not raw:
        return None
    return float(raw)


def _build_upload_state(upload_id: str) -> dict:
    meta = web_uploads.load_meta(upload_id)
    raw_name = meta.get(web_uploads.RAW_DATA_NAME + "_original")
    if not raw_name and meta.get("gemini_csv"):
        raw_name = "Gemini converted CSV"
    return {
        "raw": {
            "ready": bool(web_uploads.get_raw_data_path(upload_id)),
            "name": raw_name,
        },
        "master": {
            "ready": bool(web_uploads.get_master_path(upload_id)),
            "name": meta.get(web_uploads.MASTER_NAME + "_original"),
        },
        "secondary": {
            "ready": bool(web_uploads.get_secondary_path(upload_id)),
            "name": meta.get(web_uploads.SECONDARY_NAME + "_original"),
            "source": meta.get("secondary_source", "file"),
        },
        "gemini_csv": bool(meta.get("gemini_csv")),
    }


def _config_from_form() -> dict:
    return {
        "month": request.form.get("month", "May'26").strip(),
        "total_gmv": request.form.get("total_gmv", "").strip(),
        "threshold": request.form.get("threshold", "45") or "45",
        "raw_sheet": request.form.get("raw_sheet", "").strip() or None,
    }


def _run_pipeline(*args, **kwargs):
    """Lazy import so `python web_app.py` starts the server immediately."""
    from run_pipeline import PipelineValidationError, execute_pipeline

    try:
        return execute_pipeline(*args, **kwargs), None
    except PipelineValidationError as exc:
        return None, exc


@app.route("/health")
def health():
    return "ok", 200


def _detect_suggested_month(master_path: str) -> Optional[str]:
    try:
        wb = load_workbook_safe(master_path, data_only=True)
        suggested = detect_suggested_new_month(wb)
        wb.close()
        return suggested
    except Exception:
        return None


@app.route("/")
def index():
    upload_id = _ensure_session()
    meta = web_uploads.load_meta(upload_id)
    suggested = meta.get("suggested_month") or ""
    cfg = {
        "month": suggested or "May'26",
        "total_gmv": "",
        "threshold": "45",
        "raw_sheet": "",
    }
    last_run = meta.get("last_run")
    if last_run and not os.path.exists(last_run.get("output_path", "")):
        last_run = None
    return render_template(
        "index.html",
        uploads=_build_upload_state(upload_id),
        config=cfg,
        stages=STAGES,
        gemini_csv_ready=bool(meta.get("gemini_csv")),
        suggested_month=suggested,
        last_master_month=meta.get("last_master_month", ""),
        last_run=last_run,
    )


@app.route("/upload/raw", methods=["POST"])
def upload_raw():
    upload_id = _ensure_session()
    try:
        web_uploads.save_raw_data(upload_id, request.files.get("raw_data"))
        flash("Raw Data uploaded.", "ok")
    except Exception as exc:
        flash(str(exc), "error")
    return redirect(url_for("index"))


@app.route("/upload/master", methods=["POST"])
def upload_master():
    upload_id = _ensure_session()
    try:
        path = web_uploads.save_master(upload_id, request.files.get("master"))
        meta = web_uploads.load_meta(upload_id)
        suggested = _detect_suggested_month(path)
        if suggested:
            meta["suggested_month"] = suggested
            from shared.month_utils import scan_contiguous_headers
            from shared.anchors import SHEET_ANCHORS
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb["Raw Data"]
            a = SHEET_ANCHORS["Raw Data"]
            _, last_label = scan_contiguous_headers(ws, a["header_row"], a["series_start"])
            wb.close()
            meta["last_master_month"] = last_label
            web_uploads.save_meta(upload_id, meta)
            flash(
                f"Master uploaded. Latest month in file: {last_label}. "
                f"Suggested next month: {suggested}.",
                "ok",
            )
        else:
            flash("Master workbook uploaded.", "ok")
    except Exception as exc:
        flash(str(exc), "error")
    return redirect(url_for("index"))


@app.route("/upload/secondary", methods=["POST"])
def upload_secondary():
    upload_id = _ensure_session()
    try:
        web_uploads.save_secondary(upload_id, request.files.get("secondary"))
        meta = web_uploads.load_meta(upload_id)
        meta["secondary_source"] = "file"
        web_uploads.save_meta(upload_id, meta)
        flash("Secondary file stored (not used by pipeline yet).", "ok")
    except Exception as exc:
        flash(str(exc), "error")
    return redirect(url_for("index"))


@app.route("/upload/secondary-url", methods=["POST"])
def upload_secondary_url():
    upload_id = _ensure_session()
    try:
        web_uploads.save_secondary_url(upload_id, request.form.get("secondary_url", ""))
        flash("Secondary URL saved (not used by pipeline yet).", "ok")
    except Exception as exc:
        flash(str(exc), "error")
    return redirect(url_for("index"))


@app.route("/gemini/convert", methods=["POST"])
def gemini_convert():
    upload_id = _ensure_session()
    api_key = request.form.get("gemini_api_key", "")
    secondary = web_uploads.get_secondary_path(upload_id)
    if not secondary:
        flash("Upload a secondary file or URL first.", "error")
        return redirect(url_for("index"))

    ext = os.path.splitext(secondary)[1].lower()
    if ext not in gemini_ingest.GEMINI_CONVERT_EXTENSIONS:
        flash(
            f"Gemini convert needs image, PDF, or Excel. Got {ext!r}.",
            "error",
        )
        return redirect(url_for("index"))

    try:
        csv_text = gemini_ingest.convert_to_raw_csv(api_key, secondary)
        web_uploads.save_gemini_csv(upload_id, csv_text)
        flash("Gemini conversion OK — CSV ready as Raw Data input.", "ok")
    except Exception as exc:
        flash(f"Gemini failed: {exc}", "error")
    finally:
        # API key is form-only; never store in session or disk
        pass

    return redirect(url_for("index"))


@app.route("/run/<stage_id>", methods=["POST"])
def run_stage(stage_id: str):
    upload_id = _ensure_session()
    stage_ids = {s["id"]: s for s in STAGES}
    if stage_id not in stage_ids:
        flash(f"Unknown stage: {stage_id}", "error")
        return redirect(url_for("index"))

    master = web_uploads.get_master_path(upload_id)
    raw = web_uploads.get_raw_data_path(upload_id)
    if not master:
        flash("Upload master workbook first.", "error")
        return redirect(url_for("index"))
    if not raw:
        flash("Upload Raw Data (or convert with Gemini) first.", "error")
        return redirect(url_for("index"))

    cfg = _config_from_form()
    if not cfg["month"]:
        flash("New month label is required.", "error")
        return redirect(url_for("index"))

    try:
        total_gmv = _parse_total_gmv(cfg["total_gmv"])
        threshold = float(cfg["threshold"])
    except ValueError:
        flash("Invalid number in Total GMV or threshold.", "error")
        return redirect(url_for("index"))

    out_path = os.path.join(
        web_uploads.session_dir(upload_id),
        f"output_{stage_id}.xlsx",
    )

    try:
        result, validation_err = _run_pipeline(
            master,
            raw,
            new_month=cfg["month"],
            threshold_pp=threshold,
            raw_sheet=cfg["raw_sheet"],
            total_gmv_check=total_gmv,
            stage=stage_id,
            output_path=out_path,
        )
        if validation_err is not None:
            flash(str(validation_err), "error")
            return redirect(url_for("index"))
        out_path, warnings, info = result
    except Exception as exc:
        flash(f"Pipeline failed: {exc}", "error")
        traceback.print_exc()
        return redirect(url_for("index"))

    if warnings:
        flash("Warnings: " + " | ".join(warnings), "warn")

    download_label = stage_ids[stage_id]["download_name"]
    return send_file(
        out_path,
        as_attachment=True,
        download_name=download_label,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/process", methods=["POST"])
def process_data():
    """Single 'Process the Data' action — runs the full pipeline end to end
    and stores the result so the Download box on the page can serve it."""
    upload_id = _ensure_session()

    master = web_uploads.get_master_path(upload_id)
    raw = web_uploads.get_raw_data_path(upload_id)
    if not master:
        flash("Upload the master workbook first.", "error")
        return redirect(url_for("index"))
    if not raw:
        flash("Upload Raw Data (or convert with Gemini) first.", "error")
        return redirect(url_for("index"))

    cfg = _config_from_form()
    if not cfg["month"]:
        flash("New month label is required.", "error")
        return redirect(url_for("index"))

    try:
        total_gmv = _parse_total_gmv(cfg["total_gmv"])
        threshold = float(cfg["threshold"])
    except ValueError:
        flash("Invalid number in Total GMV or threshold.", "error")
        return redirect(url_for("index"))

    out_path = os.path.join(web_uploads.session_dir(upload_id), "output_final.xlsx")

    try:
        result, validation_err = _run_pipeline(
            master,
            raw,
            new_month=cfg["month"],
            threshold_pp=threshold,
            raw_sheet=cfg["raw_sheet"],
            total_gmv_check=total_gmv,
            stage="full_master",
            output_path=out_path,
        )
        if validation_err is not None:
            flash(str(validation_err), "error")
            return redirect(url_for("index"))
        out_path, warnings, info = result
    except Exception as exc:
        flash(f"Processing failed: {exc}", "error")
        traceback.print_exc()
        return redirect(url_for("index"))

    meta = web_uploads.load_meta(upload_id)
    meta["last_run"] = {
        "output_path": out_path,
        "download_name": f"Meesho {cfg['month']} output.xlsx",
        "month": cfg["month"],
        "categories": info["categories"],
        "unmapped": info["unmapped"],
        "warnings": warnings,
    }
    web_uploads.save_meta(upload_id, meta)

    msg = f"Processed {cfg['month']} — {info['categories']} categories written. Scroll down to download."
    if warnings:
        flash(msg, "ok")
        flash("Warnings: " + " | ".join(warnings), "warn")
    else:
        flash(msg, "ok")
    return redirect(url_for("index"))


@app.route("/download/final")
def download_final():
    upload_id = _ensure_session()
    meta = web_uploads.load_meta(upload_id)
    last_run = meta.get("last_run")
    if not last_run or not os.path.exists(last_run.get("output_path", "")):
        flash("No processed file yet — click Process the Data first.", "error")
        return redirect(url_for("index"))
    return send_file(
        last_run["output_path"],
        as_attachment=True,
        download_name=last_run.get("download_name", "Meesho output.xlsx"),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/session/clear", methods=["POST"])
def clear_session():
    upload_id = session.pop("upload_id", None)
    if upload_id:
        web_uploads.cleanup_session(upload_id)
    flash("Session uploads cleared.", "ok")
    return redirect(url_for("index"))


if __name__ == "__main__":
    port = DEFAULT_PORT
    host = DEFAULT_HOST
    url = f"http://{host}:{port}"

    print(f"\nMeesho web UI ready → {url}", flush=True)
    print("Keep this window open while using the browser.", flush=True)
    print("Press Ctrl+C to stop.\n", flush=True)

    try:
        app.run(host=host, port=port, debug=False, use_reloader=False, threaded=True)
    except OSError as exc:
        if "address already in use" in str(exc).lower() or getattr(exc, "winerror", None) == 10048:
            alt = port + 1
            print(f"Port {port} busy — trying {alt} …", flush=True)
            print(f"Open → http://{host}:{alt}\n", flush=True)
            app.run(host=host, port=alt, debug=False, use_reloader=False, threaded=True)
        else:
            print(f"\nERROR: Could not start server: {exc}", flush=True)
            print(
                "Tip: Windows sometimes blocks port 5000. "
                f"This app uses port {port} by default. "
                "Set MEESHO_WEB_PORT=9090 if needed.",
                flush=True,
            )
            input("Press Enter to exit …")
            raise
