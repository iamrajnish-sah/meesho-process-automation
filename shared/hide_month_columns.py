"""Discover month-column sections and apply user-chosen column visibility."""
from dataclasses import dataclass
from typing import Dict, List, Optional, Set, Tuple, TypedDict

from openpyxl.utils import get_column_letter as CL

from shared.month_utils import parse_month_label

SKIP_HEADER_WORDS = ("MoM", "YoY", "Deviation", "FORMULA", "Checker", "Gap", "Suggested", "Contri")

DEFAULT_VISIBLE_MONTH_LABELS = frozenset({
    "Apr'25", "May'25", "Apr'26", "May'26",
})


class HideSheetSpec(TypedDict, total=False):
    header_row: int
    block_width: int


HIDE_SHEET_CONFIG: Dict[str, HideSheetSpec] = {
    "Raw Working Sheet": {"header_row": 2, "block_width": 4},
    "Raw Data": {"header_row": 24},
    "Error Margin - Expert vs Report": {"header_row": 5},
    "FK Prelim View": {"header_row": 6},
    "New_Meesho Masterfile": {"header_row": 8},
}


@dataclass(frozen=True)
class MonthSection:
    sheet: str
    header_row: int
    section_index: int
    start_col: int
    end_col: int
    block_width: int
    months: Tuple[Tuple[str, int, int, int], ...]  # label, col_n, mon, year

    @property
    def section_key(self) -> str:
        return f"{self.sheet}|{self.section_index}|{CL(self.start_col)}"

    @property
    def label(self) -> str:
        first = self.months[0][0] if self.months else "?"
        last = self.months[-1][0] if self.months else "?"
        return (
            f"{self.sheet} — section {self.section_index + 1} "
            f"({CL(self.start_col)}…{CL(self.end_col)}: {first} → {last})"
        )


def _parse_month_header(text: str) -> Optional[Tuple[int, int, str]]:
    label = text.strip()
    if not label or label.startswith("="):
        return None
    if any(w in label for w in SKIP_HEADER_WORDS):
        return None
    core = label[:-4].strip() if label.endswith(" New") else label
    if " " in core or core.count("'") != 1:
        return None
    parsed = parse_month_label(core)
    if not parsed:
        return None
    return parsed[0], parsed[1], core


def discover_month_sections(wb) -> List[MonthSection]:
    """List every contiguous month-column run on configured sheets."""
    sections: List[MonthSection] = []
    for sheet_name, spec in HIDE_SHEET_CONFIG.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        header_row = spec["header_row"]
        block_width = spec.get("block_width", 1)
        run: List[Tuple[str, int, int, int]] = []
        run_start = 0

        def flush_run(idx: int) -> None:
            nonlocal run, run_start
            if not run:
                return
            sections.append(
                MonthSection(
                    sheet=sheet_name,
                    header_row=header_row,
                    section_index=idx,
                    start_col=run_start,
                    end_col=run[-1][1],
                    block_width=block_width,
                    months=tuple(run),
                )
            )
            run = []

        section_idx = 0
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            parsed = _parse_month_header(str(v)) if v is not None else None
            if parsed:
                if not run:
                    run_start = c
                run.append((parsed[2], c, parsed[0], parsed[1]))
            else:
                if run:
                    flush_run(section_idx)
                    section_idx += 1
        if run:
            flush_run(section_idx)
    return sections


def print_month_section_catalog(sections: List[MonthSection]) -> None:
    print("\n  [Hide] Month-column sections found (choose visible months per section):")
    for sec in sections:
        month_list = ", ".join(m[0] for m in sec.months)
        print(f"    • {sec.label}")
        print(f"      Months: {month_list}")


def prompt_visible_months(
    sections: List[MonthSection],
    *,
    interactive: bool = True,
) -> Dict[str, Set[str]]:
    """Return {section_key: {Mon'YY labels to keep visible}}."""
    if not sections:
        return {}

    print_month_section_catalog(sections)
    if not interactive:
        raise ValueError(
            "visible_months config required for non-interactive runs. "
            "Pass visible_months={section_key: [labels...]} to run_full_pipeline."
        )

    print(
        "\n  [Hide] For each section, enter month labels to KEEP visible "
        "(comma-separated Mon'YY, e.g. Nov'25,Dec'25,Apr'26,May'26)."
    )
    print("  [Hide] Press Enter to keep ALL months visible for that section.")

    config: Dict[str, Set[str]] = {}
    for sec in sections:
        available = {m[0] for m in sec.months}
        while True:
            raw = input(f"\n  Visible months for:\n    {sec.label}\n  > ").strip()
            if not raw:
                config[sec.section_key] = set(available)
                break
            chosen = {p.strip() for p in raw.split(",") if p.strip()}
            unknown = chosen - available
            if unknown:
                print(f"  ⚠ Unknown month(s) {unknown}. Available: {sorted(available)}")
                continue
            config[sec.section_key] = chosen
            break
    return config


def build_default_visible_months(
    sections: List[MonthSection],
) -> Dict[str, Set[str]]:
    """Apr'25, May'25, Apr'26, May'26 — intersected with each section's available months."""
    return build_blanket_visible_months(sections, set(DEFAULT_VISIBLE_MONTH_LABELS))


def parse_visible_month_labels(raw: str) -> Optional[Set[str]]:
    """Parse comma-separated month labels, e.g. \"Apr'25, May'25, Apr'26, May'26\"."""
    text = raw.strip()
    if not text:
        return None
    return {part.strip() for part in text.split(",") if part.strip()}


def build_blanket_visible_months(
    sections: List[MonthSection],
    labels: Set[str],
) -> Dict[str, Set[str]]:
    """Apply the same visible-month set to every discovered section."""
    config: Dict[str, Set[str]] = {}
    for sec in sections:
        available = {m[0] for m in sec.months}
        config[sec.section_key] = available & labels
    return config


def apply_column_visibility(
    wb,
    sections: List[MonthSection],
    visible_months: Dict[str, Set[str]],
    *,
    dry_run: bool = False,
) -> Dict[str, Tuple[int, int]]:
    """Hide month columns not listed in visible_months for each section."""
    stats: Dict[str, Tuple[int, int]] = {}
    print("\n  [Hide] Applying column visibility:")

    for sec in sections:
        keep = visible_months.get(sec.section_key, {m[0] for m in sec.months})
        hidden = visible = 0
        kept_labels: List[str] = []

        for label, col_n, _, _ in sec.months:
            show = label in keep
            for bc in range(col_n, col_n + sec.block_width):
                if show:
                    visible += 1
                    if label not in kept_labels:
                        kept_labels.append(label)
                else:
                    if not dry_run:
                        wb[sec.sheet].column_dimensions[CL(bc)].hidden = True
                    hidden += 1

        stats[sec.sheet] = (
            stats.get(sec.sheet, (0, 0))[0] + hidden,
            stats.get(sec.sheet, (0, 0))[1] + visible,
        )
        print(
            f"  [Hide] {sec.label}: {hidden} hidden, {visible} visible "
            f"({', '.join(kept_labels) or 'none'})"
        )

    return stats


def hide_month_columns_interactive(
    wb,
    *,
    visible_months: Optional[Dict[str, Set[str]]] = None,
    blanket_labels: Optional[Set[str]] = None,
    interactive: bool = True,
    dry_run: bool = False,
) -> Dict[str, Tuple[int, int]]:
    """Discover sections, resolve visibility config, apply hiding."""
    sections = discover_month_sections(wb)
    if not sections:
        print("  [Hide] No month sections found — skipping.")
        return {}

    if visible_months is None:
        if blanket_labels is not None:
            visible_months = build_blanket_visible_months(sections, blanket_labels)
            print(
                f"\n  [Hide] Using configured visible months: "
                f"{', '.join(sorted(blanket_labels))}"
            )
        else:
            visible_months = build_default_visible_months(sections)
            print(
                f"\n  [Hide] Using default visible months: "
                f"{', '.join(sorted(DEFAULT_VISIBLE_MONTH_LABELS))}"
            )
    else:
        print_month_section_catalog(sections)
        print("  [Hide] Using pre-supplied visible_months config.")

    return apply_column_visibility(wb, sections, visible_months, dry_run=dry_run)
