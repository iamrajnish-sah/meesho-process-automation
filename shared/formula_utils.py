"""Column arithmetic and formula shift/remap utilities."""
import re
from typing import Dict, List, Optional

from openpyxl.utils import column_index_from_string, get_column_letter

from shared.style_utils import (
    YELLOW_FILL, add_comment, copy_cell_style,
    get_row_values_across_cols, trailing_avg,
)

_CELL_RE = re.compile(r"(\$?)([A-Z]+)(\$?)(\d+)")


def col_num(col_str: str) -> int:
    return column_index_from_string(col_str)


def col_str(n: int) -> str:
    return get_column_letter(n)


def next_col(col: str, steps: int = 1) -> str:
    return col_str(col_num(col) + steps)


def shift_formula(formula: str, col_delta: int) -> str:
    """Shift RELATIVE column refs by col_delta. Absolute ($) refs unchanged."""
    if not formula:
        return formula

    def _replace(m: re.Match) -> str:
        dollar_col, letters, dollar_row, digits = m.groups()
        if dollar_col == "$":
            return m.group(0)
        new_n = col_num(letters) + col_delta
        if new_n < 1:
            return m.group(0)
        return f"{col_str(new_n)}{dollar_row}{digits}"

    return _CELL_RE.sub(_replace, formula)


def remap_formula_refs(formula: str, col_map: Dict[str, str], col_shift: int = 0) -> str:
    """Replace explicit column letters in formulas, then optionally shift relative refs.

    When col_shift is nonzero, columns substituted via col_map are protected from the
    shift step using lowercase placeholder tokens (e.g. "zzmap0zz"). shift_formula's
    [A-Z]+ regex cannot match lowercase tokens, so substituted columns are shifted
    exactly once (by the col_map substitution) rather than twice (col_map + col_shift).

    When col_shift is 0, takes a fast path identical to the previous behaviour.
    All existing callers that pass col_shift=0 (copy_column_block, the second call site
    in client_prelim.py) are therefore completely unaffected by this change.
    """
    if not formula or not isinstance(formula, str):
        return formula

    result = formula

    if not col_shift:
        # Fast path: no shifting needed — apply col_map directly, same as before.
        for old, new in sorted(col_map.items(), key=lambda x: -len(x[0])):
            for sheet_prefix in (
                "'Meesho Prelim View'!", "'Raw Data'!",
                "'Error Margin - Expert vs Report'!",
            ):
                result = result.replace(f"{sheet_prefix}{old}", f"{sheet_prefix}{new}")
                result = result.replace(f"{sheet_prefix}${old}$", f"{sheet_prefix}${new}$")
            result = re.sub(rf"(?<![A-Z])(\$?){old}(\$?\d+)", rf"\1{new}\2", result)
        return result

    # col_shift is nonzero: use placeholder tokens so col_map-substituted columns
    # survive shift_formula untouched, then restore them afterwards.
    placeholder_map: Dict[str, str] = {}  # token -> final new-column letter
    for i, (old, new) in enumerate(sorted(col_map.items(), key=lambda x: -len(x[0]))):
        token = f"zzmap{i}zz"
        placeholder_map[token] = new
        for sheet_prefix in (
            "'Meesho Prelim View'!", "'Raw Data'!",
            "'Error Margin - Expert vs Report'!",
        ):
            result = result.replace(f"{sheet_prefix}{old}", f"{sheet_prefix}{token}")
            result = result.replace(f"{sheet_prefix}${old}$", f"{sheet_prefix}${token}$")
        result = re.sub(
            rf"(?<![A-Z])(\$?){old}(\$?\d+)",
            lambda m, t=token: f"{m.group(1)}{t}{m.group(2)}",
            result,
        )

    # Shift all remaining (non-mapped) column references by col_shift.
    result = shift_formula(result, col_shift)

    # Restore placeholders to their final mapped column letters.
    for token, new in placeholder_map.items():
        result = result.replace(token, new)

    return result


def drag_column_forward(
    ws, prev_col: str, new_col: str,
    row_start: int, row_end: int,
    hardcoded_rows: Optional[List[int]] = None,
    dry_run: bool = False,
) -> int:
    """Copy formulas prev_col → new_col, shifting relative refs by +1."""
    hardcoded_rows = hardcoded_rows or []
    delta = col_num(new_col) - col_num(prev_col)
    count = 0
    for r in range(row_start, row_end + 1):
        src = ws[f"{prev_col}{r}"]
        if src.value is None:
            continue
        dst = ws[f"{new_col}{r}"]
        if r in hardcoded_rows:
            nums = []
            for cc in [col_str(col_num(prev_col) - i) for i in range(2, -1, -1)]:
                try:
                    nums.append(float(ws[f"{cc}{r}"].value))
                except (TypeError, ValueError):
                    nums.append(None)
            suggested = trailing_avg(nums, 3)
            if not dry_run and suggested is not None:
                dst.value = round(suggested, 6)
                dst.fill  = YELLOW_FILL
                copy_cell_style(src, dst)
                dst.fill  = YELLOW_FILL
                add_comment(ws, f"{new_col}{r}",
                            f"TRENDLINE SUGGESTION: trailing-3m avg = {suggested:.4f}. "
                            "Review before locking.")
                count += 1
        elif isinstance(src.value, str) and src.value.startswith("="):
            if not dry_run:
                dst.value = shift_formula(src.value, delta)
                copy_cell_style(src, dst)
                count += 1
        elif isinstance(src.value, (int, float)):
            if not dry_run:
                dst.value = src.value
                copy_cell_style(src, dst)
                dst.fill  = YELLOW_FILL
                add_comment(ws, f"{new_col}{r}",
                            "Hardcoded value copied from prior month. Review & update.")
                count += 1
    return count


def copy_column_block(
    ws, prev_col: str, new_col: str,
    row_start: int, row_end: int,
    col_map: Dict[str, str],
    header_row: Optional[int] = None,
    month_label: str = "",
    mirror_expert_col: Optional[str] = None,
    hardcoded_rows: Optional[List[int]] = None,
    dry_run: bool = False,
) -> int:
    """Copy a column block using explicit col_map remapping (no extra shift)."""
    hardcoded_rows = hardcoded_rows or []
    count = 0

    if header_row and month_label and not dry_run:
        ws[f"{new_col}{header_row}"].value = month_label
        copy_cell_style(ws[f"{prev_col}{header_row}"], ws[f"{new_col}{header_row}"])

    for r in range(row_start, row_end + 1):
        src = ws[f"{prev_col}{r}"]
        if src.value is None:
            continue
        if dry_run:
            count += 1
            continue

        dst = ws[f"{new_col}{r}"]
        val = src.value

        if r in hardcoded_rows and isinstance(val, (int, float)):
            nums = get_row_values_across_cols(ws, r, [
                col_str(col_num(prev_col) - i) for i in range(2, -1, -1)
            ])
            suggested = trailing_avg(nums, 3)
            if suggested is not None:
                dst.value = round(suggested, 6)
                dst.fill = YELLOW_FILL
                add_comment(ws, f"{new_col}{r}",
                            "TRENDLINE SUGGESTION — review before locking.")
                count += 1
                continue

        if isinstance(val, str) and val.startswith("="):
            mirrored = None
            if mirror_expert_col:
                m = re.match(r"=\$?U\$?(\d+)$", val, re.I)
                if m and prev_col == "AM":
                    mirrored = f"={mirror_expert_col}{m.group(1)}"
            dst.value = mirrored if mirrored else remap_formula_refs(val, col_map, col_shift=0)
            copy_cell_style(src, dst)
            count += 1
        elif isinstance(val, str):
            dst.value = val
            copy_cell_style(src, dst)
            count += 1
        elif isinstance(val, (int, float)):
            dst.value = val
            copy_cell_style(src, dst)
            count += 1

    return count
