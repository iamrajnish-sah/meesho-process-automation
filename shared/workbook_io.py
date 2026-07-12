"""Safe workbook loading when Excel has the file open."""
import os
import shutil
import tempfile

from openpyxl import load_workbook


def load_workbook_safe(path: str, data_only: bool = False):
    try:
        return load_workbook(path, data_only=data_only)
    except PermissionError:
        tmp = os.path.join(tempfile.gettempdir(), "meesho_wb_readcopy.xlsx")
        shutil.copy2(path, tmp)
        print("  [INFO] Master file is open in Excel — using temp copy for read.")
        return load_workbook(tmp, data_only=data_only)
