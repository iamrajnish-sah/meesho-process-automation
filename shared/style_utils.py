"""Cell style, fill, comment, and trendline helpers."""
import copy
from typing import List, Optional

from openpyxl.comments import Comment as XLComment
from openpyxl.styles import Font, PatternFill

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
RED_ANOMALY_FONT = Font(color="FF0000", bold=True)
ANOMALY_NOTE = "Please re-check again."


def copy_cell_style(src, dst) -> None:
    if src.has_style:
        dst.font          = copy.copy(src.font)
        dst.border        = copy.copy(src.border)
        dst.alignment     = copy.copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection    = copy.copy(src.protection)


def add_comment(ws, cell_ref: str, text: str) -> None:
    ws[cell_ref].comment = XLComment(text, "MeeshoBot")


def get_row_values_across_cols(ws, row: int, cols: List[str]) -> List[Optional[float]]:
    vals = []
    for col in cols:
        v = ws[f"{col}{row}"].value
        try:
            vals.append(float(v))
        except (TypeError, ValueError):
            vals.append(None)
    return vals


def trailing_avg(values: List[Optional[float]], n: int = 3) -> Optional[float]:
    non_null = [v for v in values if v is not None]
    if not non_null:
        return None
    return sum(non_null[-n:]) / len(non_null[-n:])


def mark_anomaly(ws, cell_ref: str) -> None:
    """Highlight a cell red with a re-check note. Used for MoM/YoY anomalies."""
    ws[cell_ref].font = RED_ANOMALY_FONT
    add_comment(ws, cell_ref, ANOMALY_NOTE)
