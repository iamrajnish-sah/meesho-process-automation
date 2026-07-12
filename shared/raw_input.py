"""Load user's raw category data from CSV or Excel."""
import csv
import os
import re
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook


def normalize_category_name(name) -> str:
    if name is None:
        return ""
    s = str(name).strip().lower()
    s = s.replace("&", " and ")
    s = s.replace("'", "'").replace("'", "'")
    s = s.replace("décor", "decor").replace("dcor", "decor")
    s = s.replace("stationery", "stationary")
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def parse_numeric(value) -> int:
    if value is None:
        raise ValueError("empty value")
    if isinstance(value, (int, float)):
        return int(round(value))
    s = str(value).strip().replace(",", "")
    if not s:
        raise ValueError("empty string")
    return int(float(s))


def _looks_like_header(cat, orders, gmv) -> bool:
    cat_norm = normalize_category_name(cat)
    header_words = {"category", "categories", "particular", "name", "expert", "sno", "s.no"}
    if any(w in cat_norm for w in header_words):
        return True
    try:
        parse_numeric(orders)
        parse_numeric(gmv)
        return False
    except ValueError:
        return True


def load_raw_data_excel(path: str, sheet_name: Optional[str] = None) -> Dict[str, Tuple[int, int]]:
    wb_in = load_workbook(path, data_only=True, read_only=True)
    ws_in = wb_in[sheet_name] if sheet_name else wb_in.active
    data: Dict[str, Tuple[int, int]] = {}
    for i, row in enumerate(ws_in.iter_rows(min_row=1, values_only=True)):
        if not row or len(row) < 3:
            continue
        cat, orders_raw, gmv_raw = row[0], row[1], row[2]
        if cat is None or str(cat).strip() == "":
            continue
        if i == 0 and _looks_like_header(cat, orders_raw, gmv_raw):
            print(f"  [INFO] Skipping header row: {cat!r}")
            continue
        try:
            data[str(cat).strip()] = (parse_numeric(orders_raw), parse_numeric(gmv_raw))
        except ValueError:
            print(f"  [WARN] Skipping row {i + 1}: could not read numbers for {cat!r}")
    wb_in.close()
    return data


def load_raw_data_csv(path: str) -> Dict[str, Tuple[int, int]]:
    data: Dict[str, Tuple[int, int]] = {}
    with open(path, newline="", encoding="utf-8-sig") as f:
        for i, row in enumerate(csv.reader(f)):
            if len(row) < 3:
                continue
            cat = row[0].strip()
            if not cat:
                continue
            if i == 0 and _looks_like_header(cat, row[1], row[2]):
                continue
            try:
                data[cat] = (parse_numeric(row[1]), parse_numeric(row[2]))
            except ValueError:
                print(f"  [WARN] Skipping row {i + 1}: could not read numbers for {cat!r}")
    return data


def load_raw_data(path: str, sheet_name: Optional[str] = None) -> Dict[str, Tuple[int, int]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xlsm", ".xls"):
        return load_raw_data_excel(path, sheet_name)
    return load_raw_data_csv(path)
