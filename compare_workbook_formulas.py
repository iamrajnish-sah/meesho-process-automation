#!/usr/bin/env python3
"""
compare_workbook_formulas.py — regression helper for the sheet-wise split.

Compares formula cells between two workbooks (e.g. monolith output vs modular output).
Reports any cell where the formula string differs.

Usage:
    python compare_workbook_formulas.py baseline.xlsx candidate.xlsx \\
        --sheets "Raw Data" "Meesho Prelim View"

    python compare_workbook_formulas.py baseline.xlsx candidate.xlsx  # all sheets
"""
import argparse
import sys
from typing import List, Optional, Set

from openpyxl import load_workbook


def collect_formulas(wb, sheet_names: Optional[List[str]] = None) -> dict:
    """Return {(sheet, coordinate): formula_string} for all formula cells."""
    formulas = {}
    names = sheet_names or wb.sheetnames
    for sn in names:
        if sn not in wb.sheetnames:
            print(f"  [WARN] Sheet not found: {sn!r}")
            continue
        ws = wb[sn]
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    formulas[(sn, cell.coordinate)] = v
    return formulas


def compare(baseline_path: str, candidate_path: str,
            sheet_names: Optional[List[str]] = None) -> int:
    print(f"Baseline:  {baseline_path}")
    print(f"Candidate: {candidate_path}")
    wb_base = load_workbook(baseline_path, data_only=False, read_only=True)
    wb_cand = load_workbook(candidate_path, data_only=False, read_only=True)

    base_f = collect_formulas(wb_base, sheet_names)
    cand_f = collect_formulas(wb_cand, sheet_names)
    wb_base.close()
    wb_cand.close()

    all_keys: Set = set(base_f) | set(cand_f)
    diffs = []
    only_base = []
    only_cand = []

    for key in sorted(all_keys):
        b = base_f.get(key)
        c = cand_f.get(key)
        if b is None:
            only_cand.append(key)
        elif c is None:
            only_base.append(key)
        elif b != c:
            diffs.append((key, b, c))

    print(f"\nFormula cells in baseline:  {len(base_f)}")
    print(f"Formula cells in candidate: {len(cand_f)}")
    print(f"Only in baseline:  {len(only_base)}")
    print(f"Only in candidate: {len(only_cand)}")
    print(f"Different formulas: {len(diffs)}")

    if diffs:
        print("\n── DIFFERENCES (first 30) ──")
        for (key, b, c) in diffs[:30]:
            sn, coord = key
            print(f"  {sn}!{coord}")
            print(f"    baseline:  {b[:120]}")
            print(f"    candidate: {c[:120]}")
        if len(diffs) > 30:
            print(f"  … and {len(diffs) - 30} more")

    if only_cand[:5]:
        print("\n── Only in candidate (first 5) ──")
        for key in only_cand[:5]:
            print(f"  {key[0]}!{key[1]}: {cand_f[key][:80]}")

    if not diffs and not only_base and not only_cand:
        print("\n✓ PASS — all formula cells match byte-for-byte.")
        return 0

    print("\n✗ FAIL — formulas differ.")
    return 1


def main():
    parser = argparse.ArgumentParser(description="Compare formulas between two workbooks")
    parser.add_argument("baseline", help="Reference workbook (known-good)")
    parser.add_argument("candidate", help="Workbook to verify")
    parser.add_argument("--sheets", nargs="*", default=None,
                        help="Limit comparison to these sheet names")
    args = parser.parse_args()
    sys.exit(compare(args.baseline, args.candidate, args.sheets))


if __name__ == "__main__":
    main()
